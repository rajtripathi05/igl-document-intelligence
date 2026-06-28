"""Excel generation boundary for ERP-ready Shipping Bill exports.

Independent from the Purchase Order Excel generator. Excel files are ALWAYS
generated from validated, standardized JSON, never directly from Gemini output.
The workbook contains two sheets:

    Sheet 1 - "Shipping Bill Summary"
    Sheet 2 - "Line Items"
"""

from __future__ import annotations

import io
import logging
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

_HEADER_FILL = PatternFill("solid", fgColor="0B5394")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_SECTION_FONT = Font(bold=True, color="0B5394", size=12)
_LABEL_FONT = Font(bold=True)
_THIN = Side(style="thin", color="BBBBBB")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_LINE_ITEM_COLUMNS: list[tuple[str, str]] = [
    ("line_number", "Line #"),
    ("product_description", "Product Description"),
    ("hs_code", "HS Code"),
    ("quantity", "Quantity"),
    ("unit", "Unit (UQC)"),
    ("unit_price", "Unit Price"),
    ("fob_value", "FOB Value"),
    ("gross_weight", "Gross Weight"),
    ("net_weight", "Net Weight"),
    ("number_of_packages", "Packages"),
]


class ShippingBillExcelGenerator:
    """Generate two-sheet Shipping Bill Excel workbooks from standardized JSON."""

    def build_workbook(self, shipping_bill: dict[str, Any]) -> Workbook:
        """Build a workbook from standardized Shipping Bill JSON.

        Args:
            shipping_bill: Standardized Shipping Bill data.

        Returns:
            An openpyxl Workbook populated entirely from the JSON.
        """
        workbook = Workbook()
        summary_sheet = workbook.active
        summary_sheet.title = "Shipping Bill Summary"
        self._build_summary_sheet(summary_sheet, shipping_bill)

        items_sheet = workbook.create_sheet("Line Items")
        self._build_line_items_sheet(items_sheet, shipping_bill)

        logger.info("Shipping Bill workbook built with %d sheet(s).", len(workbook.sheetnames))
        return workbook

    def generate_shipping_bill_excel(
        self, shipping_bill: dict[str, Any], output_path: Path
    ) -> Path:
        """Generate and save a Shipping Bill Excel file.

        Args:
            shipping_bill: Standardized Shipping Bill data.
            output_path: Destination path inside ``outputs/``.

        Returns:
            The path the workbook was saved to.
        """
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook = self.build_workbook(shipping_bill)
        workbook.save(output_path)
        logger.info("Saved Shipping Bill Excel workbook to %s", output_path)
        return output_path

    def to_bytes(self, shipping_bill: dict[str, Any]) -> bytes:
        """Render the workbook to bytes for in-memory download.

        Args:
            shipping_bill: Standardized Shipping Bill data.

        Returns:
            The workbook serialized as XLSX bytes.
        """
        buffer = io.BytesIO()
        self.build_workbook(shipping_bill).save(buffer)
        return buffer.getvalue()

    # ----- Sheet builders ------------------------------------------------- #

    def _build_summary_sheet(self, sheet: Worksheet, data: dict[str, Any]) -> None:
        """Populate the Shipping Bill Summary sheet."""
        sb = data.get("shipping_bill", {}) or {}
        exporter = data.get("exporter", {}) or {}
        consignee = data.get("consignee", {}) or {}
        buyer = data.get("buyer", {}) or {}
        invoice = data.get("invoice", {}) or {}
        shipment = data.get("shipment", {}) or {}
        summary = data.get("summary", {}) or {}
        drawback = data.get("drawback", {}) or {}

        row = 1
        row = self._write_section(sheet, row, "Shipping Bill")
        row = self._write_pairs(
            sheet,
            row,
            [
                ("Shipping Bill Number", sb.get("shipping_bill_number")),
                ("Shipping Bill Date", sb.get("shipping_bill_date")),
                ("Customs House", sb.get("customs_house")),
                ("Port Code", sb.get("port_code")),
                ("Port of Loading", sb.get("port_of_loading")),
                ("Port of Discharge", sb.get("port_of_discharge")),
                ("Country of Destination", sb.get("country_of_destination")),
                ("Shipping Scheme", sb.get("shipping_scheme")),
                ("Export Promotion Scheme", sb.get("export_promotion_scheme")),
                ("LEO Number", sb.get("leo_number")),
                ("LEO Date", sb.get("leo_date")),
            ],
        )
        row += 1
        row = self._write_section(sheet, row, "Exporter")
        row = self._write_pairs(
            sheet,
            row,
            [
                ("Name", exporter.get("name")),
                ("Address", exporter.get("address")),
                ("IEC Number", exporter.get("iec_number")),
                ("GSTIN", exporter.get("gstin")),
            ],
        )
        row += 1
        row = self._write_section(sheet, row, "Consignee / Buyer")
        row = self._write_pairs(
            sheet,
            row,
            [
                ("Consignee Name", consignee.get("name")),
                ("Consignee Address", consignee.get("address")),
                ("Consignee Country", consignee.get("country")),
                ("Buyer Name", buyer.get("name")),
                ("Buyer Country", buyer.get("country")),
            ],
        )
        row += 1
        row = self._write_section(sheet, row, "Invoice")
        row = self._write_pairs(
            sheet,
            row,
            [
                ("Invoice Number", invoice.get("invoice_number")),
                ("Invoice Date", invoice.get("invoice_date")),
                ("Invoice Value", invoice.get("invoice_value")),
                ("FOB Value", invoice.get("fob_value")),
                ("Freight", invoice.get("freight")),
                ("Insurance", invoice.get("insurance")),
                ("Currency", invoice.get("currency")),
                ("Exchange Rate", invoice.get("exchange_rate")),
            ],
        )
        row += 1
        row = self._write_section(sheet, row, "Shipment")
        row = self._write_pairs(
            sheet,
            row,
            [
                ("Container Number", shipment.get("container_number")),
                ("Seal Number", shipment.get("seal_number")),
                ("Vehicle Number", shipment.get("vehicle_number")),
                ("Total Packages", shipment.get("total_packages")),
                ("Total Gross Weight", shipment.get("total_gross_weight")),
                ("Total Net Weight", shipment.get("total_net_weight")),
            ],
        )
        row += 1
        row = self._write_section(sheet, row, "Summary")
        row = self._write_pairs(
            sheet,
            row,
            [
                ("Total Invoice Value", summary.get("total_invoice_value")),
                ("Total FOB Value", summary.get("total_fob_value")),
                ("Total Freight", summary.get("total_freight")),
                ("Total Insurance", summary.get("total_insurance")),
                ("Currency", summary.get("currency")),
            ],
        )
        row += 1
        row = self._write_section(sheet, row, "Drawback")
        self._write_pairs(
            sheet,
            row,
            [
                ("Drawback Claimed", drawback.get("drawback_claimed")),
                ("Drawback Amount", drawback.get("drawback_amount")),
                ("Drawback Details", drawback.get("drawback_details")),
            ],
        )

        sheet.column_dimensions["A"].width = 24
        sheet.column_dimensions["B"].width = 70

    def _build_line_items_sheet(self, sheet: Worksheet, data: dict[str, Any]) -> None:
        """Populate the Line Items sheet."""
        items = data.get("items") or []

        for col_index, (_, header) in enumerate(_LINE_ITEM_COLUMNS, start=1):
            cell = sheet.cell(row=1, column=col_index, value=header)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _BORDER

        for row_index, item in enumerate(items, start=2):
            item = item or {}
            for col_index, (key, _) in enumerate(_LINE_ITEM_COLUMNS, start=1):
                cell = sheet.cell(row=row_index, column=col_index, value=item.get(key))
                cell.border = _BORDER
                cell.alignment = Alignment(
                    vertical="top", wrap_text=key == "product_description"
                )

        self._autosize(sheet, max_width=50)
        sheet.freeze_panes = "A2"

    # ----- Low-level writers --------------------------------------------- #

    def _write_section(self, sheet: Worksheet, row: int, title: str) -> int:
        """Write a section header row and return the next row index."""
        cell = sheet.cell(row=row, column=1, value=title)
        cell.font = _SECTION_FONT
        return row + 1

    def _write_pairs(
        self, sheet: Worksheet, row: int, pairs: list[tuple[str, Any]]
    ) -> int:
        """Write label/value pairs and return the next row index."""
        for label, value in pairs:
            label_cell = sheet.cell(row=row, column=1, value=label)
            label_cell.font = _LABEL_FONT
            label_cell.alignment = Alignment(vertical="top")
            value_cell = sheet.cell(row=row, column=2, value=value)
            value_cell.alignment = Alignment(vertical="top", wrap_text=True)
            row += 1
        return row

    def _autosize(self, sheet: Worksheet, max_width: int = 60) -> None:
        """Approximate column auto-sizing based on cell content length."""
        widths: dict[int, int] = {}
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                length = len(str(cell.value))
                widths[cell.column] = max(widths.get(cell.column, 10), length + 2)
        for column, width in widths.items():
            sheet.column_dimensions[get_column_letter(column)].width = min(
                width, max_width
            )
