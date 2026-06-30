"""Consolidated one-row-per-document Excel registers.

Business users receive a flat register per document type: ONE workbook, ONE
worksheet, and **one row per uploaded PDF**, built from each processor's
declarative export mapping (``ProcessorSpec.export``). When the processor
declares a styled template (the uploaded business workbook, e.g. ``Book8.xlsx``
or ``tsmp (2).XLSX``), the template's header styling, fonts, borders, column
widths and merged cells are preserved and data rows are written beneath the
header — matching columns by **header name** so the mapping is robust to column
position. A batch spanning multiple document types yields one workbook per type,
bundled into a ZIP.
"""

from __future__ import annotations

import io
import logging
import re
import zipfile
from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from processors.base import BaseProcessor
from processors.spec import ExportSpec, ProcessorSpec
from utils.file_handler import slugify

logger = logging.getLogger(__name__)

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_THIN = Side(style="thin", color="BBBBBB")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _norm(header: Any) -> str:
    """Normalize a header for tolerant matching (case/space/punctuation-insensitive)."""
    return re.sub(r"[^a-z0-9]", "", str(header or "").lower())


def _excel_safe(value: Any) -> Any:
    """Coerce a value into something openpyxl can write."""
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def build_register(processor: BaseProcessor, docs: list) -> bytes:
    """Build a one-row-per-document register workbook for one processor.

    Args:
        processor: The processor whose export mapping defines the columns.
        docs: Document states of this processor's type (each contributes a row).

    Returns:
        XLSX bytes for the register.
    """
    spec = processor.spec
    export = spec.export
    rows = [processor.build_row(d.data or {}) for d in docs if getattr(d, "data", None)]

    template_path = processor.template_path() if hasattr(processor, "template_path") else None
    if export and export.template and template_path and Path(template_path).is_file():
        try:
            return _build_from_template(Path(template_path), export, rows)
        except Exception:  # noqa: BLE001 - fall back to a fresh sheet on any template issue
            logger.exception("Template register build failed; using a fresh sheet.")
    return _build_fresh(spec, export, rows)


def _build_from_template(
    template_path: Path, export: ExportSpec, rows: list[dict[str, Any]]
) -> bytes:
    """Write rows into a copy of the styled template, matching columns by header."""
    workbook = load_workbook(template_path)
    sheet = workbook[export.sheet] if export.sheet in workbook.sheetnames else workbook.active

    header_map = _header_index(sheet, export.header_row)
    target: dict[str, int] = {}
    next_col = sheet.max_column
    for column in export.columns:
        col_index = header_map.get(_norm(column.header))
        if col_index is None:
            next_col += 1
            col_index = next_col
            header_cell = sheet.cell(row=export.header_row, column=col_index, value=column.header)
            header_cell.fill = _HEADER_FILL
            header_cell.font = _HEADER_FONT
            header_cell.border = _BORDER
        target[column.header] = col_index

    ref_styles = _capture_row_styles(sheet, export.start_row, target.values())
    _clear_data_rows(sheet, export.start_row)

    for offset, row in enumerate(rows):
        excel_row = export.start_row + offset
        for column in export.columns:
            col_index = target[column.header]
            cell = sheet.cell(row=excel_row, column=col_index, value=_excel_safe(row.get(column.header)))
            _apply_style(cell, ref_styles.get(col_index))

    buffer = io.BytesIO()
    workbook.save(buffer)
    logger.info("Built register from template %s with %d row(s).", template_path.name, len(rows))
    return buffer.getvalue()


def _build_fresh(spec: ProcessorSpec, export: ExportSpec | None, rows: list[dict[str, Any]]) -> bytes:
    """Build a clean register workbook when no template is declared."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = (export.sheet if export else "Register")[:31]

    headers = (
        [c.header for c in export.columns]
        if export and export.columns
        else (list(rows[0].keys()) if rows else [])
    )
    for col_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_index, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _BORDER

    for row_offset, row in enumerate(rows, start=2):
        for col_index, header in enumerate(headers, start=1):
            cell = sheet.cell(row=row_offset, column=col_index, value=_excel_safe(row.get(header)))
            cell.border = _BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    _autosize(sheet)
    sheet.freeze_panes = "A2"

    buffer = io.BytesIO()
    workbook.save(buffer)
    logger.info("Built fresh register '%s' with %d row(s).", sheet.title, len(rows))
    return buffer.getvalue()


def build_registers(docs: list) -> dict[str, bytes]:
    """Build one register workbook per document type present in ``docs``.

    Args:
        docs: Processed document states.

    Returns:
        ``{filename: xlsx_bytes}`` — one entry per document type.
    """
    groups: dict[str, list] = {}
    processors: dict[str, BaseProcessor] = {}
    for doc in docs:
        processor = getattr(doc, "processor", None)
        if processor is None or getattr(doc, "data", None) is None:
            continue
        key = processor.spec.use_case_key
        groups.setdefault(key, []).append(doc)
        processors[key] = processor

    registers: dict[str, bytes] = {}
    for key, group in groups.items():
        processor = processors[key]
        label = processor.spec.business_process or processor.spec.document_type or key
        registers[f"{slugify(label)}_register.xlsx"] = build_register(processor, group)
    return registers


def registers_zip(docs: list) -> bytes:
    """Bundle per-type registers into a single ZIP (for multi-type batches)."""
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as archive:
        for name, data in build_registers(docs).items():
            archive.writestr(name, data)
    return buffer.getvalue()


# ----- Template helpers -------------------------------------------------- #


def _header_index(sheet: Worksheet, header_row: int) -> dict[str, int]:
    """Map normalized header text → column index (first occurrence wins)."""
    mapping: dict[str, int] = {}
    for col in range(1, sheet.max_column + 1):
        key = _norm(sheet.cell(row=header_row, column=col).value)
        if key and key not in mapping:
            mapping[key] = col
    return mapping


def _capture_row_styles(sheet: Worksheet, row: int, columns: Any) -> dict[int, dict[str, Any]]:
    """Snapshot per-column cell styles from a reference data row for reuse."""
    styles: dict[int, dict[str, Any]] = {}
    if row > sheet.max_row:
        return styles
    for col in columns:
        cell = sheet.cell(row=row, column=col)
        styles[col] = {
            "font": copy(cell.font),
            "border": copy(cell.border),
            "alignment": copy(cell.alignment),
            "number_format": cell.number_format,
        }
    return styles


def _apply_style(cell: Any, style: dict[str, Any] | None) -> None:
    """Apply a captured style snapshot to a cell (no-op when absent)."""
    if not style:
        return
    cell.font = copy(style["font"])
    cell.border = copy(style["border"])
    cell.alignment = copy(style["alignment"])
    cell.number_format = style["number_format"]


def _clear_data_rows(sheet: Worksheet, start_row: int) -> None:
    """Clear values in all rows from ``start_row`` down (styling preserved)."""
    for row in range(start_row, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            sheet.cell(row=row, column=col).value = None


def _autosize(sheet: Worksheet, max_width: int = 48) -> None:
    """Approximate column auto-sizing for a fresh register."""
    widths: dict[int, int] = {}
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            widths[cell.column] = max(widths.get(cell.column, 10), len(str(cell.value)) + 2)
    for column, width in widths.items():
        sheet.column_dimensions[get_column_letter(column)].width = min(width, max_width)
