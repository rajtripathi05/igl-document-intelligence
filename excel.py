"""Excel generation boundary for ERP-ready Purchase Order exports.

Excel files are ALWAYS generated from validated, standardized JSON, never
directly from Gemini output (see CLAUDE.md). The workbook contains two sheets:

    Sheet 1 - "Purchase Order Summary"
    Sheet 2 - "Line Items"
"""

from __future__ import annotations

import io
import logging
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_SECTION_FONT = Font(bold=True, color="1F4E78", size=12)
_LABEL_FONT = Font(bold=True)
_THIN = Side(style="thin", color="BBBBBB")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_LINE_ITEM_COLUMNS: list[tuple[str, str]] = [
    ("line_number", "Line #"),
    ("material_code", "Material Code"),
    ("description", "Description"),
    ("hsn_code", "HSN Code"),
    ("quantity", "Quantity"),
    ("unit", "Unit"),
    ("unit_price", "Unit Price"),
    ("taxable_amount", "Taxable Amount"),
    ("gst_type", "GST Type"),
    ("gst_percent", "GST %"),
    ("gst_amount", "GST Amount"),
    ("total_amount", "Total Amount"),
]


class ExcelGenerator:
    """Generate two-sheet Purchase Order Excel workbooks from standardized JSON."""

    def build_workbook(self, purchase_order: dict[str, Any]) -> Workbook:
        """Build a workbook from standardized Purchase Order JSON.

        Args:
            purchase_order: Standardized Purchase Order data.

        Returns:
            An openpyxl Workbook populated entirely from the JSON.
        """
        workbook = Workbook()
        summary_sheet = workbook.active
        summary_sheet.title = "Purchase Order Summary"
        self._build_summary_sheet(summary_sheet, purchase_order)

        items_sheet = workbook.create_sheet("Line Items")
        self._build_line_items_sheet(items_sheet, purchase_order)

        logger.info("Workbook built with %d sheet(s).", len(workbook.sheetnames))
        return workbook

    def generate_purchase_order_excel(
        self, purchase_order: dict[str, Any], output_path: Path
    ) -> Path:
        """Generate and save a Purchase Order Excel file.

        Args:
            purchase_order: Standardized Purchase Order data.
            output_path: Destination path inside ``outputs/``.

        Returns:
            The path the workbook was saved to.
        """
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook = self.build_workbook(purchase_order)
        workbook.save(output_path)
        logger.info("Saved Excel workbook to %s", output_path)
        return output_path

    def to_bytes(self, purchase_order: dict[str, Any]) -> bytes:
        """Render the workbook to bytes for in-memory download.

        Args:
            purchase_order: Standardized Purchase Order data.

        Returns:
            The workbook serialized as XLSX bytes.
        """
        buffer = io.BytesIO()
        self.build_workbook(purchase_order).save(buffer)
        return buffer.getvalue()

    # ----- Sheet builders ------------------------------------------------- #

    def _build_summary_sheet(
        self, sheet: Worksheet, data: dict[str, Any]
    ) -> None:
        """Populate the Purchase Order Summary sheet."""
        po = data.get("purchase_order", {}) or {}
        buyer = data.get("buyer", {}) or {}
        supplier = data.get("supplier", {}) or {}
        summary = data.get("summary", {}) or {}

        row = 1
        row = self._write_section(sheet, row, "Purchase Order")
        row = self._write_pairs(
            sheet,
            row,
            [
                ("PO Number", po.get("po_number")),
                ("PO Date", po.get("po_date")),
                ("Supplier Reference", po.get("supplier_reference")),
                ("Payment Terms", po.get("payment_terms")),
                ("Delivery Date", po.get("delivery_date")),
                ("Currency", po.get("currency")),
                ("Nature of Supply", po.get("nature_of_supply")),
            ],
        )
        row += 1
        row = self._write_section(sheet, row, "Buyer")
        row = self._write_pairs(
            sheet,
            row,
            [
                ("Company Name", buyer.get("company_name")),
                ("Address", buyer.get("address")),
                ("GST Number", buyer.get("gst_number")),
                ("PAN Number", buyer.get("pan_number")),
            ],
        )
        row += 1
        row = self._write_section(sheet, row, "Supplier")
        row = self._write_pairs(
            sheet,
            row,
            [
                ("Company Name", supplier.get("company_name")),
                ("Address", supplier.get("address")),
                ("GST Number", supplier.get("gst_number")),
                ("PAN Number", supplier.get("pan_number")),
            ],
        )
        row += 1
        row = self._write_section(sheet, row, "Summary")
        self._write_pairs(
            sheet,
            row,
            [
                ("Subtotal", summary.get("subtotal")),
                ("CGST", summary.get("cgst")),
                ("SGST", summary.get("sgst")),
                ("IGST", summary.get("igst")),
                ("Grand Total", summary.get("grand_total")),
                ("Amount in Words", summary.get("amount_in_words")),
            ],
        )

        sheet.column_dimensions["A"].width = 22
        sheet.column_dimensions["B"].width = 70

    def _build_line_items_sheet(
        self, sheet: Worksheet, data: dict[str, Any]
    ) -> None:
        """Populate the Line Items sheet."""
        items = data.get("items") or []

        for col_index, (_, header) in enumerate(_LINE_ITEM_COLUMNS, start=1):
            cell = sheet.cell(row=1, column=col_index, value=header)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _BORDER

        for row_index, item in enumerate(items, start=2):
            item = item or {}
            for col_index, (key, _) in enumerate(_LINE_ITEM_COLUMNS, start=1):
                cell = sheet.cell(row=row_index, column=col_index, value=item.get(key))
                cell.border = _BORDER
                cell.alignment = Alignment(vertical="top", wrap_text=key == "description")

        self._autosize(sheet, max_width=50)
        sheet.freeze_panes = "A2"

    # ----- Low-level writers --------------------------------------------- #

    def _write_section(self, sheet: Worksheet, row: int, title: str) -> int:
        """Write a section header row and return the next row index."""
        cell = sheet.cell(row=row, column=1, value=title)
        cell.font = _SECTION_FONT
        return row + 1

    def _write_pairs(
        self, sheet: Worksheet, row: int, pairs: list[tuple[str, Any]]
    ) -> int:
        """Write label/value pairs and return the next row index."""
        for label, value in pairs:
            label_cell = sheet.cell(row=row, column=1, value=label)
            label_cell.font = _LABEL_FONT
            label_cell.alignment = Alignment(vertical="top")
            value_cell = sheet.cell(row=row, column=2, value=value)
            value_cell.alignment = Alignment(vertical="top", wrap_text=True)
            row += 1
        return row

    def _autosize(self, sheet: Worksheet, max_width: int = 60) -> None:
        """Approximate column auto-sizing based on cell content length."""
        widths: dict[int, int] = {}
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                length = len(str(cell.value))
                widths[cell.column] = max(widths.get(cell.column, 10), length + 2)
        for column, width in widths.items():
            sheet.column_dimensions[get_column_letter(column)].width = min(
                width, max_width
            )
