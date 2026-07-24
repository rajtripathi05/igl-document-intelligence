"""Log Book engine — many mobile-photo pages -> one register row per batch.

This is the deterministic assembly + export half of the ``engine: "logbook"``
processor. AI is used ONLY to read one page image at a time (via the shared,
provider-agnostic gateway through the processor's :class:`GeminiClient`); this
module never talks to a provider or a key. It:

1. Extracts each uploaded page independently (cached by content, cost-tracked).
2. Classifies each page as a printed grid MAIN page (Stirred Reactor Log Book or
   General Batch Card) or a lined REMARKS narrative page.
3. Assembles batches by pairing each main page with its remarks page using
   content join-keys (batch no, date, unloading/filling times, total kg) with a
   weak upload-order tiebreak — robust when hundreds of images arrive flat and
   out of order.
4. Writes ONE combined Excel register — one row per batch — in the uploaded
   Book11 layout (grouped 3-row header) plus a Remarks column, companion
   RM-material-name columns, and dynamically-added columns for any QC row or
   specification parameter that does not fit the fixed template.

The module is Streamlit-free so it can be unit-tested and driven from a
background worker. ``app.py`` wraps :func:`run` with the upload UI.
"""

from __future__ import annotations

import io
import logging
import re
from typing import Any, Callable, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

ProgressCallback = Callable[[int, int, str], None]

# ─────────────────────────────────────────────────────────────────────────────
# Page-type constants
# ─────────────────────────────────────────────────────────────────────────────
MAIN_TYPES = {"reactor_main", "blending_main", "main"}
REMARKS_TYPES = {"remarks", "remark", "shift_process_activities"}

# ─────────────────────────────────────────────────────────────────────────────
# Raw-material name → fixed INPUT column (LA / KOH / EO). Everything else flows
# into the RM-4..RM-12 slots in encounter order, with its name kept in a
# companion "RM-n Material" column so identity is never lost.
# ─────────────────────────────────────────────────────────────────────────────
_RM_CANON: list[tuple[str, set[str]]] = [
    ("LA", {"la", "laa", "laurylalcohol", "lauryl", "la1", "fattyalcohol"}),
    ("KOH", {"koh", "potassiumhydroxide", "kotf"}),
    ("EO", {"eo", "ethyleneoxide", "eqo"}),
]

# Specification parameter → fixed Book11 "Complete Analysis" column. Anything not
# listed here is appended as its own added column ("<param>"), so any product's
# parameters (MEA/DEA/TEA, B.Value, Iodine, Free PO, Viscosity, …) are preserved.
_SPEC_CANON: list[tuple[str, set[str]]] = [
    ("App", {"app", "appearance", "appear"}),
    ("Meout", {"meout", "moist", "moust", "moisture", "moisture%", "meust",
               "mout", "mst", "mc", "moistcontent"}),
    ("Alpha", {"alpha", "apha", "aphacolour", "aphacolor", "colourapha", "hazen"}),
    ("CP", {"cp", "cloudpoint"}),
    ("OH", {"oh", "ohvalue", "ohv", "hydroxyl", "hydroxylvalue", "ohno", "ohnumber"}),
    ("Free EO", {"freeeo", "freeo", "freeethyleneoxide", "feo"}),
    ("1,4 Dioxane", {"14dioxane", "dioxane", "onefourdioxane", "1,4dioxane"}),
    ("pEch", {"pech", "peg", "pch", "pegvalue"}),
]

# Canonical display name for a Complete-Analysis / Specification parameter so the
# same reading with handwriting/OCR spelling variants collapses to ONE column
# (PH/Ph/pH -> "pH"; Vis/Visc./Viscosity -> "Viscosity"; Moist/Meist/Maist ->
# "Moisture"). Unknown parameters are kept as written.
_SPEC_SYNONYMS: dict[str, str] = {
    "app": "App", "appearance": "App", "appear": "App", "abb": "App", "appe": "App",
    "moisture": "Moisture", "moist": "Moisture", "meist": "Moisture", "maist": "Moisture",
    "meout": "Moisture", "meust": "Moisture", "maint": "Moisture", "mons": "Moisture",
    "mout": "Moisture", "mst": "Moisture", "moisture": "Moisture",
    "apha": "APHA", "alpha": "APHA", "adha": "APHA", "aphacolour": "APHA",
    "hazen": "APHA", "colourapha": "APHA",
    "cp": "CP", "cip": "CP", "cloudpoint": "CP",
    "oh": "OH", "ohvalue": "OH", "ohv": "OH", "ohno": "OH", "hydroxyl": "OH",
    "withbaseoh": "With Base OH", "withoutbaseoh": "Without Base OH",
    "freeeo": "Free EO", "freeo": "Free EO", "feo": "Free EO", "freeera": "Free EO",
    "14dioxane": "1,4 Dioxane", "dioxane": "1,4 Dioxane", "14diox": "1,4 Dioxane",
    "diox": "1,4 Dioxane", "onefourdioxane": "1,4 Dioxane",
    "peg": "PEG", "pech": "PEG", "pch": "PEG", "pegvalue": "PEG",
    "ph": "pH",
    "bvalue": "B.Value", "bval": "B.Value", "basevalue": "B.Value", "bv": "B.Value",
    "i2": "Iodine (I2)", "iodine": "Iodine (I2)", "iodinevalue": "Iodine (I2)", "iv": "Iodine (I2)",
    "vis": "Viscosity", "visc": "Viscosity", "viscosity": "Viscosity", "visco": "Viscosity",
    "viscosity25c": "Viscosity", "viscosityat25c": "Viscosity",
    "freepo": "Free PO", "freepropyleneoxide": "Free PO",
    "acidvalue": "Acid Value", "acidval": "Acid Value", "av": "Acid Value",
    "sap": "SAP", "sapvalue": "SAP", "saponification": "SAP",
    "calltest": "Call Test", "coldtest": "Cold Test",
    "spgr": "Sp. Gr", "specificgravity": "Sp. Gr", "spgravity": "Sp. Gr",
    "spgr25c": "Sp. Gr", "spgrat25c": "Sp. Gr",
    "ir": "IR", "particle": "Particle", "colour": "Colour", "color": "Colour",
}

#: Preferred column order for specification parameters (present ones first).
_SPEC_ORDER = [
    "App", "Moisture", "APHA", "CP", "OH", "Free EO", "1,4 Dioxane", "PEG", "pH",
    "Viscosity", "B.Value", "Iodine (I2)", "Free PO", "With Base OH",
    "Without Base OH", "Acid Value", "SAP", "Sp. Gr", "IR", "Call Test",
    "Cold Test", "Particle", "Colour",
]


def _norm_spec(name: Any) -> str:
    """Canonical display name for a specification parameter (folds OCR variants)."""
    raw = str(name or "").strip()
    key = re.sub(r"[^a-z0-9]", "", raw.lower())
    if not key:
        return raw or "Spec"
    if key in _SPEC_SYNONYMS:
        return _SPEC_SYNONYMS[key]
    for suffix in ("value", "no", "content", "25c", "at25c"):
        if key.endswith(suffix) and key[: -len(suffix)] in _SPEC_SYNONYMS:
            return _SPEC_SYNONYMS[key[: -len(suffix)]]
    return raw


# Raw-material name synonyms -> canonical display name, so each material that is
# NOT LA/KOH/EO gets ONE self-documenting column (e.g. every "A.Acid"/"Acetic
# Acid" reading lands together). Unknown materials are kept exactly as written
# (e.g. "ENLITE 275T"), so you always see what raw material is actually there.
_MATERIAL_SYNONYMS: dict[str, str] = {
    "po": "PO", "propyleneoxide": "PO",
    "aacid": "A.Acid", "aceticacid": "A.Acid", "acid": "A.Acid", "aa": "A.Acid",
    "dmwater": "DM Water", "demineralisedwater": "DM Water", "dmw": "DM Water",
    "water": "Water",
    "ammoniasol": "Ammonia Sol", "ammonia": "Ammonia Sol",
    "ammoniasolution": "Ammonia Sol", "ammoniasoln": "Ammonia Sol",
    "soyabeanoil": "Soyabean Oil", "soybeanoil": "Soyabean Oil", "soyaoil": "Soyabean Oil",
    "meoh": "MEOH", "methanol": "MEOH",
    "hcl": "HCl", "hydrochloricacid": "HCl",
    "mea": "MEA", "dea": "DEA", "tea": "TEA", "koh": "KOH",
}


def _norm_material(name: Any) -> str:
    """Canonical display name for a raw material (folds common OCR variants)."""
    raw = str(name or "").strip()
    key = re.sub(r"[^a-z0-9]", "", raw.lower())
    if not key:
        return raw or "RM"
    if key in _MATERIAL_SYNONYMS:
        return _MATERIAL_SYNONYMS[key]
    # Unknown material: collapse punctuation/extra spaces so "ENLITE-275T",
    # "ENLITE 275T" and "ENLITE -275T" land in ONE column. (Genuine letter/digit
    # misreads like "275T" vs "2757" remain distinct — those need the stronger
    # model, LOGBOOK_MODEL=strong.)
    return re.sub(r"[^A-Za-z0-9]+", " ", raw).strip() or raw


_QC_MIN_SLOTS = 5  # Book11 provides five in-process QC slots.

_ADDED_GROUP = "Log Book (added)"
_SPEC_GROUP = "Complete Analysis / Specification"


# ═════════════════════════════════════════════════════════════════════════════
# Normalization helpers
# ═════════════════════════════════════════════════════════════════════════════
def _clean_key(text: Any) -> str:
    """Lowercase alphanumerics only — for tolerant name matching."""
    return re.sub(r"[^a-z0-9]", "", str(text or "").lower())


def _to_number(value: Any) -> float | int | str | None:
    """Best-effort numeric coercion; preserves notation like '<10 ppm', '51-57'."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip()
    if not text:
        return None
    # A clean number possibly with a unit suffix (kg, %, etc.).
    m = re.fullmatch(r"[-+]?\d{1,3}(?:,\d{3})*(?:\.\d+)?\s*[A-Za-z%°/]*\.?", text)
    if m:
        num = re.sub(r"[^0-9.\-+]", "", text.split()[0].replace(",", ""))
        try:
            f = float(num)
            return int(f) if f.is_integer() else f
        except ValueError:
            return text
    return text  # keep original notation (ranges, <, >, ppm, mm, N/D, …)


def _num_for_match(value: Any) -> float | None:
    """Extract a plain float for numeric comparison, or None."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    m = re.search(r"[-+]?\d+(?:\.\d+)?", str(value).replace(",", ""))
    return float(m.group()) if m else None


def _norm_time(value: Any) -> str | None:
    """Normalize a clock time to 'HH:MM' (24h). Accepts 12:02, 12.02, 1202, 8:5."""
    if value is None:
        return None
    text = str(value).strip().lower()
    text = re.sub(r"\s*(hrs?|hr|h|am|pm|hours?)\.?$", "", text)
    m = re.search(r"(\d{1,2})\s*[:.\-]\s*(\d{2})", text)
    if not m:
        m2 = re.fullmatch(r"(\d{1,2})(\d{2})", re.sub(r"\D", "", text))
        if not m2:
            return None
        hh, mm = int(m2.group(1)), int(m2.group(2))
    else:
        hh, mm = int(m.group(1)), int(m.group(2))
    if hh > 23 or mm > 59:
        return None
    return f"{hh:02d}:{mm:02d}"


def _norm_date(value: Any) -> tuple[int, int] | None:
    """Return (day, month) from many written forms, ignoring year. None if absent."""
    if value is None:
        return None
    text = str(value)
    m = re.search(r"\b(\d{1,2})\s*[./\-]\s*(\d{1,2})\s*[./\-]\s*(\d{2,4})\b", text)
    if not m:
        m = re.search(r"\b(\d{1,2})\s*[./\-]\s*(\d{1,2})\b", text)
        if not m:
            return None
    day, month = int(m.group(1)), int(m.group(2))
    if 1 <= day <= 31 and 1 <= month <= 12:
        return (day, month)
    if 1 <= month <= 31 and 1 <= day <= 12:  # swapped
        return (month, day)
    return None


def _iso_date(value: Any) -> str | None:
    """Return an ISO 'YYYY-MM-DD' when a full dd/mm/yy(yy) is confidently present."""
    if value is None:
        return None
    m = re.search(r"\b(\d{1,2})\s*[./\-]\s*(\d{1,2})\s*[./\-]\s*(\d{2,4})\b", str(value))
    if not m:
        return None
    day, month, year = int(m.group(1)), int(m.group(2)), int(m.group(3))
    if day > 31 and month <= 31:
        day, month = month, day
    if year < 100:
        year += 2000
    if not (1 <= day <= 31 and 1 <= month <= 12):
        return None
    try:
        from datetime import date

        return date(year, month, day).isoformat()
    except ValueError:
        return None


# ═════════════════════════════════════════════════════════════════════════════
# Page extraction (AI, per image, cached)
# ═════════════════════════════════════════════════════════════════════════════
def _prepare_parts(file_bytes: bytes, mime_type: str) -> list[tuple[bytes, str]]:
    """OCR-preprocess a page into model-ready parts; fall back to the raw image."""
    try:
        import preprocess

        parts = [(p.data, p.mime_type) for p in preprocess.prepare(file_bytes, mime_type)]
        if parts:
            return parts
    except Exception:  # noqa: BLE001 - never fail extraction on preprocessing
        logger.debug("Preprocess failed; using raw bytes.", exc_info=True)
    return [(file_bytes, mime_type)]


def _extract_page(
    client: Any,
    filename: str,
    file_bytes: bytes,
    mime_type: str,
    *,
    prompt_version: str,
    use_cache: bool,
) -> tuple[dict[str, Any], dict[str, Any]]:
    """Extract one page via the client, using the shared extraction cache."""
    data: dict[str, Any] | None = None
    scores: dict[str, Any] = {}
    key = None
    if use_cache:
        try:
            import cache

            key = cache.make_key(file_bytes, "logbook", prompt_version)
            cached = cache.load(key)
            if cached is not None:
                return cached.get("data", {}) or {}, cached.get("ai_scores", {}) or {}
        except Exception:  # noqa: BLE001
            logger.debug("Cache lookup failed for %s.", filename, exc_info=True)

    parts = _prepare_parts(file_bytes, mime_type)
    # Default: the cheaper DEFAULT_MODEL (flash). Log Book pages are handwritten,
    # so if accuracy is not enough you can opt in to the STRONGER model
    # (RETRY_MODEL, e.g. gemini-2.5-pro) by setting LOGBOOK_MODEL=strong in .env —
    # no code change needed.
    import os

    use_strong = (os.getenv("LOGBOOK_MODEL", "flash").strip().lower() == "strong")
    data, scores = client.extract_with_confidence(parts, use_retry_model=use_strong)
    data = data or {}
    if use_cache and key is not None:
        try:
            import cache

            cache.store(key, data, scores)
        except Exception:  # noqa: BLE001
            logger.debug("Cache store failed for %s.", filename, exc_info=True)
    return data, scores


def _page_confidence(scores: dict[str, Any]) -> int | None:
    """Average of the AI per-field confidences on a page (0-100), or None."""
    vals = [v for v in (scores or {}).values() if isinstance(v, (int, float))]
    return round(sum(vals) / len(vals)) if vals else None


# ═════════════════════════════════════════════════════════════════════════════
# Batch assembly (pair each main page with its remarks page)
# ═════════════════════════════════════════════════════════════════════════════
class _Page:
    """A single extracted page with its classification and join signals."""

    def __init__(self, index: int, filename: str, group: str, data: dict[str, Any],
                 scores: dict[str, Any]) -> None:
        self.index = index
        self.filename = filename
        self.group = group  # source group / folder hint (product), if any
        self.data = data or {}
        self.scores = scores or {}
        self.confidence = _page_confidence(scores)
        self.kind = self._classify()

    def _classify(self) -> str:
        pt = _clean_key(self.data.get("page_type"))
        if pt in {_clean_key(t) for t in MAIN_TYPES}:
            return "main"
        if pt in {_clean_key(t) for t in REMARKS_TYPES}:
            return "remarks"
        # Fall back to content shape when page_type is missing/unknown.
        if self.data.get("batch_no") or self.data.get("raw_materials"):
            return "main"
        if (self.data.get("remarks_text") or "").strip():
            return "remarks"
        return "unknown"

    # --- join signals (present on both a main page and its remarks page) --- #
    @property
    def _jk(self) -> dict[str, Any]:
        return self.data.get("join_keys") or {}

    def date(self) -> tuple[int, int] | None:
        return _norm_date(self.data.get("date")) or _norm_date(self._jk.get("date")) \
            or _norm_date(self.data.get("remarks_text"))

    def batch_hint(self) -> str:
        return _clean_key(self.data.get("batch_no") or self._jk.get("batch_hint"))

    def product_hint(self) -> str:
        return _clean_key(self.data.get("product_name") or self._jk.get("product_hint")
                          or self.group)

    def total_kg(self) -> float | None:
        ud = self.data.get("unloading_details") or {}
        return _num_for_match(ud.get("total_filled_kg") or self._jk.get("total_filled_kg"))

    def unl_finish(self) -> str | None:
        ps = (self.data.get("process_steps") or {}).get("unloading") or {}
        return _norm_time(ps.get("finish") or self._jk.get("unloading_finish"))

    def unl_start(self) -> str | None:
        ps = (self.data.get("process_steps") or {}).get("unloading") or {}
        return _norm_time(ps.get("start") or self._jk.get("unloading_start"))

    def la_start(self) -> str | None:
        jk = _norm_time(self._jk.get("la_charge_start"))
        if jk:
            return jk
        for rm in self.data.get("raw_materials") or []:
            if isinstance(rm, dict) and _clean_key(rm.get("name")).startswith("la"):
                return _norm_time(rm.get("start"))
        return None


class _Batch:
    """A main page plus any remarks pages paired to it."""

    def __init__(self, main: _Page | None) -> None:
        self.main = main
        self.remarks: list[_Page] = []


def _pair_score(main: _Page, rem: _Page) -> float:
    """Score how strongly a remarks page belongs to a main page."""
    s = 0.0
    if main.batch_hint() and rem.batch_hint() and main.batch_hint() == rem.batch_hint():
        s += 8
    mk, rk = main.total_kg(), rem.total_kg()
    if mk is not None and rk is not None and abs(mk - rk) <= 1.0:
        s += 6
    if main.unl_finish() and rem.unl_finish() and main.unl_finish() == rem.unl_finish():
        s += 5
    if main.la_start() and rem.la_start() and main.la_start() == rem.la_start():
        s += 4
    if main.unl_start() and rem.unl_start() and main.unl_start() == rem.unl_start():
        s += 3
    if main.date() and rem.date() and main.date() == rem.date():
        s += 4
    mp, rp = main.product_hint(), rem.product_hint()
    if mp and rp and (mp == rp or mp in rp or rp in mp):
        s += 3
    # Weak upload-order tiebreak: a remarks page usually follows its main page.
    gap = rem.index - main.index
    if 0 < gap <= 2:
        s += 1.5
    elif abs(gap) <= 4:
        s += 0.5
    return s


def _assemble_batches(pages: list[_Page]) -> tuple[list[_Batch], list[str]]:
    """Group pages into batches, pairing remarks pages to their main page."""
    warnings: list[str] = []
    mains = [p for p in pages if p.kind == "main"]
    rem_pages = [p for p in pages if p.kind == "remarks"]
    for p in pages:
        if p.kind == "unknown":
            warnings.append(f"Could not classify page '{p.filename}' — skipped.")

    batches = [_Batch(m) for m in mains]
    by_main = {id(m): b for m, b in zip(mains, batches)}

    if not mains:
        # No grid pages at all: emit each remarks page as its own row so nothing
        # is lost (rare — e.g. only narrative pages were uploaded).
        for r in rem_pages:
            b = _Batch(None)
            b.remarks.append(r)
            batches.append(b)
        if rem_pages:
            warnings.append("No main grid pages found; emitted remarks-only rows.")
        return batches, warnings

    for r in rem_pages:
        scored = sorted(mains, key=lambda m: _pair_score(m, r), reverse=True)
        best = scored[0]
        if _pair_score(best, r) >= 3:
            by_main[id(best)].remarks.append(r)
        else:
            # Fallback: nearest preceding main page, else nearest overall.
            preceding = [m for m in mains if m.index <= r.index]
            target = (max(preceding, key=lambda m: m.index) if preceding
                      else min(mains, key=lambda m: abs(m.index - r.index)))
            by_main[id(target)].remarks.append(r)
            warnings.append(
                f"Remarks page '{r.filename}' paired by position (weak match)."
            )
    return batches, warnings


# ═════════════════════════════════════════════════════════════════════════════
# Row building
# ═════════════════════════════════════════════════════════════════════════════
def _match_canon(name: Any, table: Iterable[tuple[str, set[str]]]) -> str | None:
    key = _clean_key(name)
    if not key:
        return None
    for canon, aliases in table:
        if key == _clean_key(canon) or key in aliases:
            return canon
    # startswith tolerance (e.g. "lauryl alcohol (fb-9001)")
    for canon, aliases in table:
        if any(key.startswith(a) for a in aliases | {_clean_key(canon)}):
            return canon
    return None


def _material_value(rm: dict[str, Any]) -> Any:
    return _to_number(rm.get("actual_qty") if rm.get("actual_qty") not in (None, "")
                      else rm.get("recipe_qty"))


def _batch_to_row(batch: _Batch) -> dict[str, Any]:
    """Reduce a batch (main page + remarks pages) to a flat register row dict."""
    main = batch.main.data if batch.main else {}
    row: dict[str, Any] = {}

    # --- identity ------------------------------------------------------- #
    row["reactor"] = main.get("reactor")
    row["date"] = _iso_date(main.get("date")) or main.get("date")
    row["batch_no"] = main.get("batch_no")
    # Q2 decision: Product Name = handwritten real product; Previous Reactor/Line
    # = the struck-through printed product name.
    row["product_name"] = main.get("product_name") or (batch.main.group if batch.main else None)
    row["previous_reactor_line"] = (
        main.get("printed_product_struck") or main.get("previous_reactor_line")
    )

    # --- raw materials: LA/KOH/EO by name, and EVERY other raw material as its
    #     own column named after the material (so you see exactly what RM is in
    #     each product, instead of generic RM-4/RM-6 slots) ------------------ #
    for canon, _ in _RM_CANON:
        row[f"rm:{canon}"] = None
    mats: dict[str, Any] = {}
    for rm in main.get("raw_materials") or []:
        if not isinstance(rm, dict) or not (rm.get("name") or rm.get("actual_qty")):
            continue
        canon = _match_canon(rm.get("name"), _RM_CANON)
        val = _material_value(rm)
        if canon and row.get(f"rm:{canon}") in (None, ""):
            row[f"rm:{canon}"] = val
            continue
        # any other material (or a duplicate LA/KOH/EO reading) -> its own column
        name = _norm_material(rm.get("name"))
        if not name:
            continue
        if name in mats and mats.get(name) not in (None, ""):
            k = 2
            while f"{name} ({k})" in mats:
                k += 1
            mats[f"{name} ({k})"] = val
        else:
            mats[name] = val
    row["_mats"] = mats

    # --- process steps -------------------------------------------------- #
    ps = main.get("process_steps") or {}
    for step in ("inertization", "dehydration", "reaction_eo", "reaction_po",
                 "cooking", "stripping", "cooling", "unloading"):
        seg = ps.get(step) or {}
        row[f"ps:{step}.start"] = _norm_time(seg.get("start")) or seg.get("start")
        row[f"ps:{step}.finish"] = _norm_time(seg.get("finish")) or seg.get("finish")

    # --- QC analysis (all rows; overflow columns added later) ----------- #
    qc = [q for q in (main.get("qc_analysis") or []) if isinstance(q, dict)
          and any(q.get(k) not in (None, "") for k in ("time", "parameter", "result"))]
    row["_qc"] = qc

    # --- complete analysis / specification (normalized param names) ------ #
    specs: dict[str, Any] = {}
    for ca in main.get("complete_analysis") or []:
        if not isinstance(ca, dict):
            continue
        param = ca.get("parameter")
        result = ca.get("result") if ca.get("result") not in (None, "") else ca.get("specs")
        if not param and result in (None, ""):
            continue
        name = _norm_spec(param)
        val = _to_number(result)
        if name in specs and specs.get(name) not in (None, ""):
            k = 2  # a second distinct reading of the same parameter
            while f"{name} ({k})" in specs:
                k += 1
            specs[f"{name} ({k})"] = val
        else:
            specs[name] = val
    row["_specx"] = specs

    # --- tail identity & added block ------------------------------------ #
    ud = main.get("unloading_details") or {}
    row["field"] = None
    row["shift_incharge_name"] = main.get("shift_incharge_name")
    row["shift_incharge_sign"] = main.get("shift_incharge_sign")
    row["packing"] = ud.get("packing_breakup")
    row["total_kg"] = _to_number(ud.get("total_filled_kg"))
    row["tank_no"] = ud.get("tank_no")
    row["panel"] = main.get("panel_incharge")
    row["form_type"] = main.get("form_type") or (batch.main.data.get("page_type") if batch.main else None)

    # Remarks: concatenate every paired remarks-page narrative (in page order).
    narratives = []
    for r in sorted(batch.remarks, key=lambda p: p.index):
        txt = (r.data.get("remarks_text") or "").strip()
        if txt:
            narratives.append(txt)
    row["remarks"] = "\n\n".join(narratives) if narratives else None

    # provenance / QA
    srcs = []
    if batch.main:
        srcs.append(batch.main.filename)
    srcs += [r.filename for r in batch.remarks]
    row["sources"] = ", ".join(srcs)
    row["pages"] = len(srcs)
    confs = [p.confidence for p in ([batch.main] if batch.main else []) + batch.remarks
             if p and p.confidence is not None]
    row["confidence"] = round(sum(confs) / len(confs)) if confs else None

    # merge any page-level extra_fields
    ex = row.get("extra", {})
    if isinstance(main.get("extra_fields"), dict):
        for k, v in main["extra_fields"].items():
            ex[str(k)] = v
    row["extra"] = ex or None
    return row


# ═════════════════════════════════════════════════════════════════════════════
# Column model + workbook
# ═════════════════════════════════════════════════════════════════════════════
def _build_columns(rows: list[dict[str, Any]]) -> list[dict[str, str]]:
    """Compute the ordered register columns (fixed + dynamic) for these rows."""
    qc_slots = max(_QC_MIN_SLOTS, max((len(r.get("_qc") or []) for r in rows), default=0))
    # union of specification parameters (normalized) — preferred order first,
    # then any others in first-seen order.
    seen_specs: list[str] = []
    for r in rows:
        for p in (r.get("_specx") or {}):
            if p not in seen_specs:
                seen_specs.append(p)
    specx_params = ([p for p in _SPEC_ORDER if p in seen_specs]
                    + [p for p in seen_specs if p not in _SPEC_ORDER])
    # union of every raw material seen beyond LA/KOH/EO, in first-seen order —
    # each becomes its own named INPUT column.
    mat_names: list[str] = []
    for r in rows:
        for m in (r.get("_mats") or {}):
            if m not in mat_names:
                mat_names.append(m)

    cols: list[dict[str, str]] = []

    def add(group, sub, name, key):
        cols.append({"group": group or "", "sub": sub or "", "name": name, "key": key})

    add(None, None, "Reactor", "reactor")
    add(None, None, "Date", "date")
    add(None, None, "Batch No.", "batch_no")
    add(None, None, "Product Name", "product_name")
    add(None, None, "Previous Reactor/Line", "previous_reactor_line")

    add("INPUT", "Input-1", "LA", "rm:LA")
    add("INPUT", "Input-2", "KOH", "rm:KOH")
    add("INPUT", "Input-3", "EO", "rm:EO")
    # one column per actual raw material used (auto-grows with the products)
    for i, m in enumerate(mat_names, start=4):
        add("INPUT", f"Input-{i}", m, f"mat:{m}")

    steps = [("Inertization", "inertization"), ("Dehydration", "dehydration"),
             ("Reaction EO", "reaction_eo"), ("Reaction PO", "reaction_po"),
             ("Cooking", "cooking"), ("Stripping", "stripping"),
             ("Cooling", "cooling"), ("Unloading", "unloading")]
    for label, key in steps:
        add("Process Steps", None, f"{label} Start", f"ps:{key}.start")
        add("Process Steps", None, f"{label} Finish", f"ps:{key}.finish")

    for i in range(1, qc_slots + 1):
        add("QC analysis", None, f"{i} time", f"qc:{i}:time")
        add("QC analysis", None, f"{i} parameter", f"qc:{i}:parameter")
        add("QC analysis", None, f"{i} result", f"qc:{i}:result")

    for p in specx_params:
        add(_SPEC_GROUP, None, p, f"specx:{p}")

    add(None, None, "Field", "field")
    add(None, None, "Shift Incharge Name", "shift_incharge_name")
    add(None, None, "Shift Incharge Sign", "shift_incharge_sign")

    # Added, log-book-specific block (the new Remarks column lives here).
    add(_ADDED_GROUP, None, "Remarks", "remarks")
    add(_ADDED_GROUP, None, "Packing / Filling", "packing")
    add(_ADDED_GROUP, None, "Total Filled (kg)", "total_kg")
    add(_ADDED_GROUP, None, "Tank No.", "tank_no")
    add(_ADDED_GROUP, None, "Panel Incharge", "panel")
    add(_ADDED_GROUP, None, "Form Type", "form_type")
    add(_ADDED_GROUP, None, "Source Image(s)", "sources")
    add(_ADDED_GROUP, None, "Pages", "pages")
    add(_ADDED_GROUP, None, "Confidence %", "confidence")
    add(_ADDED_GROUP, None, "Extra Fields", "extra")
    return cols


def _cell_value(row: dict[str, Any], key: str) -> Any:
    """Resolve a column key against a row (handles qc:/specx: dynamic keys)."""
    if key.startswith("qc:"):
        _, idx, field = key.split(":")
        qc = row.get("_qc") or []
        i = int(idx) - 1
        if 0 <= i < len(qc):
            return qc[i].get(field)
        return None
    if key.startswith("specx:"):
        return (row.get("_specx") or {}).get(key.split(":", 1)[1])
    if key.startswith("mat:"):
        return (row.get("_mats") or {}).get(key.split(":", 1)[1])
    val = row.get(key)
    if isinstance(val, dict):
        return "; ".join(f"{k}={v}" for k, v in val.items()) if val else None
    return val


_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_GROUP_FILL = PatternFill("solid", fgColor="2E75B6")
_SUB_FILL = PatternFill("solid", fgColor="9DC3E6")
_ADDED_FILL = PatternFill("solid", fgColor="548235")
_WHITE = Font(bold=True, color="FFFFFF", size=10)
_DARK = Font(bold=True, color="1F3864", size=10)
_THIN = Side(style="thin", color="BBBBBB")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _build_workbook(rows: list[dict[str, Any]], columns: list[dict[str, str]]) -> bytes:
    """Write the combined register: grouped 3-row header + one row per batch."""
    return _write_register(columns, rows, lambda row, col: _cell_value(row, col["key"]))


def _write_register(
    columns: list[dict[str, str]],
    rows: list[dict[str, Any]],
    value_fn: Callable[[dict[str, Any], dict[str, str]], Any],
) -> bytes:
    """Render the grouped 3-row header + one row per record; values via ``value_fn``."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Log Book Register"

    # Row 1: group headers (merged across each contiguous group span).
    # Row 2: sub headers (Input-1..12).  Row 3: column names.  Data from row 4.
    for c, col in enumerate(columns, start=1):
        g = ws.cell(row=1, column=c, value=col["group"] or None)
        s = ws.cell(row=2, column=c, value=col["sub"] or None)
        n = ws.cell(row=3, column=c, value=col["name"])
        n.fill = _ADDED_FILL if col["group"] == _ADDED_GROUP else _HEADER_FILL
        n.font = _WHITE
        n.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        n.border = _BORDER
        if g.value:
            g.fill = _GROUP_FILL
            g.font = _WHITE
            g.alignment = Alignment(horizontal="center", vertical="center")
        if s.value:
            s.fill = _SUB_FILL
            s.font = _DARK
            s.alignment = Alignment(horizontal="center", vertical="center")

    _merge_groups(ws, columns)

    for r, row in enumerate(rows, start=4):
        for c, col in enumerate(columns, start=1):
            v = value_fn(row, col)
            cell = ws.cell(row=r, column=c, value=_excel_safe(v))
            cell.alignment = Alignment(vertical="top", wrap_text=(col["key"] == "remarks"))
            cell.border = _BORDER

    for c, col in enumerate(columns, start=1):
        letter = get_column_letter(c)
        if col["key"] == "remarks":
            ws.column_dimensions[letter].width = 60
        elif col["key"] in ("sources", "packing", "extra"):
            ws.column_dimensions[letter].width = 26
        elif col["key"] in ("product_name", "previous_reactor_line", "batch_no"):
            ws.column_dimensions[letter].width = 18
        else:
            ws.column_dimensions[letter].width = 12
    ws.freeze_panes = "A4"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_workbook_from_display(
    column_defs: list[dict[str, str]], display_rows: list[dict[str, Any]]
) -> bytes:
    """Rebuild the styled register from (possibly edited) display rows.

    ``display_rows`` are keyed by column display name (as returned in ``run``'s
    ``rows``); ``column_defs`` is ``run``'s ``column_defs``. Lets the reviewer's
    edits in the UI flow straight back into the downloadable register.
    """
    return _write_register(column_defs, display_rows,
                           lambda row, col: row.get(col["name"]))


def _merge_groups(ws, columns: list[dict[str, str]]) -> None:
    """Merge contiguous, identical, non-empty group labels on header row 1."""
    i = 0
    n = len(columns)
    while i < n:
        g = columns[i]["group"]
        j = i
        while j + 1 < n and columns[j + 1]["group"] == g:
            j += 1
        if g and j > i:
            ws.merge_cells(start_row=1, start_column=i + 1, end_row=1, end_column=j + 1)
        i = j + 1


def _excel_safe(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


# ═════════════════════════════════════════════════════════════════════════════
# Public entry point
# ═════════════════════════════════════════════════════════════════════════════
def run(
    files: list[tuple[str, bytes]],
    *,
    client: Any,
    spec: Any = None,
    groups: dict[str, str] | None = None,
    progress: ProgressCallback | None = None,
    use_cache: bool = True,
    record_cost: bool = True,
) -> dict[str, Any]:
    """Extract, assemble, and export a batch of log-book page images.

    Args:
        files: ``(filename, raw_bytes)`` for every uploaded page image/PDF.
        client: A ``GeminiClient`` bound to the logbook prompts + schema.
        spec: The processor spec (for cost tagging); optional.
        groups: Optional ``{filename: product_hint}`` (e.g. folder name) to aid
            product/remarks pairing when the caller knows it.
        progress: ``progress(done, total, filename)`` callback.
        use_cache: Reuse the extraction cache (skips AI on repeat images).
        record_cost: Record token cost per page in the platform cost dashboard.

    Returns:
        ``{columns, rows, records, excel_bytes, stats, warnings}``.
    """
    from utils.file_handler import guess_mime_type

    prompt_version = getattr(spec, "prompt_version", "v1") if spec else "v1"
    groups = groups or {}
    total = len(files)
    pages: list[_Page] = []
    warnings: list[str] = []
    model_name = ""
    try:
        import config

        model_name = config.settings.default_model
    except Exception:  # noqa: BLE001
        model_name = ""

    for i, (filename, file_bytes) in enumerate(files):
        if progress:
            progress(i, total, filename)
        try:
            mime = guess_mime_type(filename)
        except Exception:  # noqa: BLE001 - default to jpeg for odd extensions
            mime = "image/jpeg"
        try:
            data, scores = _extract_page(
                client, filename, file_bytes, mime,
                prompt_version=prompt_version, use_cache=use_cache,
            )
        except Exception as exc:  # noqa: BLE001 - one bad page must not stop the batch
            logger.exception("Extraction failed for %s.", filename)
            warnings.append(f"Failed to read '{filename}': {exc}")
            continue

        # cost tracking (best-effort)
        if record_cost and spec is not None:
            try:
                import cost

                usage = getattr(client, "last_usage", {}) or {}
                cost.record(spec.use_case_key, spec.department_key, model_name or "logbook",
                            usage, model_role=cost.ROLE_DEFAULT, proc_ms=0, is_retry=False)
            except Exception:  # noqa: BLE001
                logger.debug("Cost record failed for %s.", filename, exc_info=True)

        pages.append(_Page(i, filename, groups.get(filename, ""), data, scores))

    if progress:
        progress(total, total, "assembling")

    batches, assemble_warnings = _assemble_batches(pages)
    warnings += assemble_warnings

    # Order rows by (product, date, batch no) for a stable, readable register.
    records = [_batch_to_row(b) for b in batches]

    def _sort_key(r: dict[str, Any]):
        return (str(r.get("product_name") or "~"), str(r.get("date") or ""),
                str(r.get("batch_no") or ""))

    records.sort(key=_sort_key)

    columns = _build_columns(records) if records else []
    excel_bytes = _build_workbook(records, columns) if records else b""

    # Flat preview rows (display name -> value), excluding internal keys.
    preview_rows = []
    for r in records:
        prow = {}
        for col in columns:
            prow[col["name"]] = _cell_value(r, col["key"])
        preview_rows.append(prow)

    main_count = sum(1 for p in pages if p.kind == "main")
    rem_count = sum(1 for p in pages if p.kind == "remarks")
    products = sorted({str(r.get("product_name")) for r in records if r.get("product_name")})
    stats = {
        "images": total,
        "pages_read": len(pages),
        "main_pages": main_count,
        "remarks_pages": rem_count,
        "batches": len(records),
        "products": len(products),
        "product_names": products,
        "model": model_name,
        "columns": len(columns),
    }
    return {
        "columns": [c["name"] for c in columns],
        "column_defs": columns,
        "rows": preview_rows,
        "records": records,
        "excel_bytes": excel_bytes,
        "stats": stats,
        "warnings": warnings,
    }
