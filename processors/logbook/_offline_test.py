"""Offline pipeline test for the logbook engine (no network / no AI).

Feeds hand-crafted page JSON (mirroring the real sample images) through the
assembly + export pipeline via a mock client, then reloads the register to check
the batch pairing, field mapping, dynamic columns, and Excel structure.

Run:  python processors/logbook/_offline_test.py
"""
from __future__ import annotations

import importlib.util
import sys
from pathlib import Path

HERE = Path(__file__).resolve().parent
ROOT = HERE.parents[1]
sys.path.insert(0, str(ROOT))

spec = importlib.util.spec_from_file_location("logbook_engine", HERE / "logbook_engine.py")
eng = importlib.util.module_from_spec(spec)
spec.loader.exec_module(eng)


class MockSpec:
    use_case_key = "logbook"
    department_key = "plant_operations"
    prompt_version = "v1"


class MockClient:
    """Returns canned page JSON in call order, ignoring the image bytes."""

    def __init__(self, pages):
        self._pages = pages
        self._i = 0

    def extract_with_confidence(self, parts):
        data = self._pages[self._i]
        self._i += 1
        scores = {k: 95 for k in ("batch_no", "reactor", "product_name")}
        return data, scores


def M(**kw):
    return kw


# --- ECARE 243: 3 batches (main) + 3 remarks, deliberately shuffled order ---
ecare_0317 = M(page_type="reactor_main", form_type="stirred_reactor_log_book",
    reactor="R-850", date="10/06/26", batch_no="NS2606060317",
    product_name="ECARE 243", printed_product_struck="ORCA-B",
    raw_materials=[
        {"slot": 1, "name": "LA", "recipe_qty": 4998, "actual_qty": 4998, "start": "08:00", "finish": "08:30"},
        {"slot": 3, "name": "KOH", "recipe_qty": 7.5, "actual_qty": 7.5, "start": "08:35", "finish": "08:45"},
        {"slot": 4, "name": "EO", "recipe_qty": 3550, "actual_qty": 3550},
    ],
    process_steps={"inertization": {"start": "08:45", "finish": "08:55"},
        "dehydration": {"start": "09:00", "finish": "12:50"},
        "reaction_eo": {"start": "13:00", "finish": "14:00"}, "reaction_po": {},
        "cooking": {"start": "14:00", "finish": "15:30"}, "stripping": {"start": "15:40", "finish": "17:10"},
        "cooling": {"start": "17:10", "finish": "17:40"}, "unloading": {"start": "18:50", "finish": "20:40"}},
    qc_analysis=[{"time": "09:45", "parameter": "m/c", "result": "0.08"},
        {"time": "11:15", "parameter": "m/c", "result": "0.05"},
        {"time": "12:15", "parameter": "m/c", "result": "0.03"},
        {"time": "14:40", "parameter": "C.P", "result": "55"}],
    complete_analysis=[{"parameter": "App", "specs": None, "result": "C"},
        {"parameter": "PH", "specs": "5.5-7.5", "result": "6.8"},
        {"parameter": "Moust", "specs": "0.5", "result": "0.07"},
        {"parameter": "Apha", "specs": "50max", "result": "20"},
        {"parameter": "C.P", "specs": "51-57", "result": "55"},
        {"parameter": "OH", "specs": "165-171", "result": "168.54"},
        {"parameter": "Free EO", "specs": "1.0max", "result": "<10 ppm"},
        {"parameter": "1,4 dioxane", "specs": "1.0max", "result": "<1 ppm"},
        {"parameter": "PEG", "specs": None, "result": "0.58"}],
    unloading_details={"total_filled_kg": 8100, "packing_breakup": "09 x 900 = 8100"},
    join_keys={"date": "10/06/26", "la_charge_start": "08:00", "unloading_start": "18:50",
               "unloading_finish": "20:40", "total_filled_kg": 8100})

ecare_0318 = M(page_type="reactor_main", reactor="R-850", date="10/06/2026", batch_no="NS2606060318",
    product_name="ECARE 243", printed_product_struck="ORCA-B",
    raw_materials=[{"name": "LA", "actual_qty": 4998, "start": "21:00", "finish": "21:30"},
        {"name": "KOH", "actual_qty": 7.5, "start": "21:30", "finish": "21:50"},
        {"name": "E.O", "recipe_qty": 3550, "actual_qty": 3550},
        {"name": "A.Acid", "actual_qty": 6.8}],
    process_steps={"unloading": {"start": "10:15", "finish": "11:50"}},
    qc_analysis=[{"time": "23:30", "parameter": "Moist", "result": "0.04"}],
    complete_analysis=[{"parameter": "App", "result": "C"}, {"parameter": "OH", "result": "168.3"}],
    unloading_details={"total_filled_kg": 9042, "packing_breakup": "10 x 900 = 9000 + 01 x 42 = 9042"},
    join_keys={"date": "10/06/26", "unloading_finish": "11:50", "total_filled_kg": 9042})

ecare_0319 = M(page_type="reactor_main", reactor="R-850", date="11.6.26", batch_no="NS2606060319",
    product_name="ECARE 243", printed_product_struck="ORCA-B",
    raw_materials=[{"name": "L.A.", "actual_qty": 4998, "start": "12:02", "finish": "12:38"},
        {"name": "KOH", "actual_qty": 7.5}, {"name": "A.Acid", "actual_qty": 6.8}],
    process_steps={"unloading": {"start": "07:00", "finish": "10:00"}},
    unloading_details={"total_filled_kg": 8100},
    join_keys={"date": "11.6.26", "la_charge_start": "12:02", "unloading_start": "07:00",
               "unloading_finish": "10:00", "total_filled_kg": 8100})

rem_0317 = M(page_type="remarks", remarks_text="1st shift 10/06/26\n# LA charged done.\n# KOH charging done @ 7.5 kg.\n# EO reaction done\nCooking & stripping completed @17:10hrs. Cooling done @17:40hrs. Filling start @18:50 & completed @20:40. 09 IBC full & one loose. Total 10 IBC. 09 x 900 = 8100 kg",
    join_keys={"date": "10/06/26", "unloading_start": "18:50", "unloading_finish": "20:40", "total_filled_kg": 8100})
rem_0318 = M(page_type="remarks", remarks_text="LA charging start @21:50... Dehydration completed 0:30. Filling over at 11:50hr. 10 x 900 = 9000 + 01 x 42 = 9042 kg",
    join_keys={"date": "10/06/26", "unloading_finish": "11:50", "total_filled_kg": 9042})
rem_0319 = M(page_type="remarks", remarks_text="11.6.26 LA charging start at 12:02h... Filling start at 07:00h. 09 x 900 = 8100 kg",
    join_keys={"date": "11.6.26", "la_charge_start": "12:02", "unloading_start": "07:00", "total_filled_kg": 8100})

# --- COZMIN: different specs (MEA/DEA/TEA) + struck ORCA-B -> handwritten MEA ---
cozmin = M(page_type="reactor_main", reactor="R-850", date="14/10/25", batch_no="NS2510050509",
    product_name="MEA", printed_product_struck="ORCA-B", previous_reactor_line=None,
    raw_materials=[{"name": "Ammonia Sol", "actual_qty": 7500, "start": "16:00", "finish": "17:32"},
        {"name": "E.O", "actual_qty": 1050, "start": "09:22", "finish": "11:30"}],
    process_steps={"cooking": {"start": "11:30", "finish": "13:30"}, "unloading": {"start": "17:40", "finish": "19:40"}},
    qc_analysis=[{"time": "12:35", "parameter": "MEA", "result": "47.64"},
        {"time": None, "parameter": "DEA", "result": "44.40"}, {"time": None, "parameter": "TEA", "result": "0.19"}],
    complete_analysis=[{"parameter": "MEA", "result": "46.83"}, {"parameter": "DEA", "result": "45.89"},
        {"parameter": "TEA", "result": "0.25"}, {"parameter": "Moisture", "result": "64.65"}],
    unloading_details={},
    join_keys={"date": "14/10/25"})

# --- Blending batch card (different form) ---
blending = M(page_type="blending_main", form_type="blending_batch_card", reactor="A-601",
    date="11/3/26", batch_no="NS2603030159", product_name="FLOSOL S018 ME", printed_product_struck=None,
    raw_materials=[{"name": "Soyabean oil", "recipe_qty": 6451, "actual_qty": 6451, "start": "18:40", "finish": "00:30"},
        {"name": "MEOH", "recipe_qty": 1190, "actual_qty": 1190}, {"name": "KOH", "recipe_qty": 35.5, "actual_qty": 35.5}],
    process_steps={}, qc_analysis=[{"time": "07:30", "parameter": "A.V", "result": "N/D"}, {"time": "12:00", "parameter": "PH", "result": "6.6"}],
    complete_analysis=[{"parameter": "App", "specs": "Particle/C", "result": "C"}, {"parameter": "PH", "specs": "4-6.5", "result": "6.3"},
        {"parameter": "Acid Value", "specs": "0.5max", "result": "0.24"}, {"parameter": "SAP", "specs": "180-202", "result": "185.2"}],
    unloading_details={}, extra_fields={"Cold test": "OK"},
    join_keys={"date": "11/3/26"})

# Shuffled upload order (like WhatsApp timestamp order): mains and remarks mixed.
files = [
    ("img_00.jpg", b"a"), ("img_01.jpg", b"b"), ("img_02.jpg", b"c"), ("img_03.jpg", b"d"),
    ("img_04.jpg", b"e"), ("img_05.jpg", b"f"), ("img_06.jpg", b"g"), ("img_07.jpg", b"h"),
]
pages = [ecare_0319, ecare_0317, rem_0317, rem_0318, ecare_0318, rem_0319, cozmin, blending]

client = MockClient(pages)
result = eng.run(files, client=client, spec=MockSpec(), use_cache=False, record_cost=False,
                 groups={"img_06.jpg": "COZMIN 060 INT 1"})

st = result["stats"]
print("STATS:", st)
print("WARNINGS:", result["warnings"])
print(f"\nAssembled {len(result['records'])} batch rows:")
for r in result["records"]:
    print(f"  • {r.get('product_name'):<16} batch={r.get('batch_no'):<14} "
          f"prevRL={r.get('previous_reactor_line')!s:<8} date={r.get('date')} "
          f"LA={r.get('rm:LA')} KOH={r.get('rm:KOH')} EO={r.get('rm:EO')} "
          f"RM4={r.get('rm:4')}({r.get('rmname:4')}) "
          f"unl={r.get('ps:unloading.start')}-{r.get('ps:unloading.finish')} "
          f"App={r.get('spec:App')} OH={r.get('spec:OH')} "
          f"specx={list((r.get('_specx') or {}).keys())} "
          f"pages={r.get('pages')} rem={'Y' if r.get('remarks') else 'N'}")

# Save + reload the register to verify structure.
out = ROOT / "outputs" / "logbook_offline_register.xlsx"
out.parent.mkdir(parents=True, exist_ok=True)
out.write_bytes(result["excel_bytes"])
print(f"\nSaved register -> {out}  ({len(result['excel_bytes'])} bytes)")

import openpyxl
wb = openpyxl.load_workbook(out)
ws = wb.active
print(f"Sheet '{ws.title}' dims={ws.dimensions} cols={ws.max_column} rows={ws.max_row}")
from openpyxl.utils import get_column_letter
hdr = []
for c in range(1, ws.max_column + 1):
    g = ws.cell(1, c).value; n = ws.cell(3, c).value
    hdr.append(f"{n}" + (f"[{g}]" if g else ""))
print("HEADERS:", " | ".join(hdr))
