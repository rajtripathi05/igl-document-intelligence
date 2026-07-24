"""RA Posting — deterministic tabular engine (v2: multi-sheet, multi-block, dedup).

Reads SBI / IDBI bank statements (Excel .xlsx/.xls/.xlsb or CSV), keeps only
CREDIT (CR) entries, and produces a customer-wise summary:

    Date | Customer Name | Amount | Document No. | Mode | Bank Name

Why v2 — the real compiled workbooks (e.g. "suman bank statements.xlsb")
broke v1 in three ways, all fixed here:

1. **Multiple sheets per workbook** (one per bank: 'IDBI', 'SBI'). v1 read only
   the first sheet. v2 reads EVERY sheet of every uploaded file.
2. **Many pasted statement blocks per sheet** — the same account is downloaded
   repeatedly and pasted below the previous download, each block with its own
   header row and an overlapping date range. v1 parsed only the first header,
   so the same transaction appeared 4-5 times. v2 detects every header row,
   re-maps columns per block, and de-duplicates: same canonical timestamp +
   narration + reference + amount = the same bank posting (the running balance
   is deliberately NOT part of the identity — a retroactive posting shifts all
   later balances between downloads and would fake uniqueness).
3. **Messy narrations** — v1 returned reference fragments as names ("2",
   "2601270023"). v2 understands the observed SBI/IDBI narration formats
   (see ``_heuristic_name``) and, for SBI, also reads the counterparty name
   from the 'Ref No./Cheque No.' column ("TRANSFER FROM <acct> / <NAME>").

Output workbook: the two business title lines, then the summary table —
customers grouped A-Z, one row per credit, a "<name> Total" subtotal per
customer and a GRAND TOTAL. **Each bank gets its own sheet**; a combined
"All Banks" sheet is added when more than one bank is present.

Hybrid design (unchanged): everything above is deterministic Python. The
shared AI gateway, when available, only cleans/normalises customer names;
if it is unavailable the summary is still produced with rule-based names.

Entry point (called by ``FolderProcessor.run_tabular``):
    run(files: list[tuple[str, bytes]], ai_gateway=None) -> dict
"""

from __future__ import annotations

import io
import json
import logging
import re
from collections import Counter
from datetime import datetime, timedelta
from typing import Any

logger = logging.getLogger(__name__)

#: Output columns, in order (mirrors the business summary template).
SUMMARY_COLUMNS = ["Date", "Customer Name", "Amount", "Document No.", "Mode", "Bank Name"]

#: The two fixed title lines above the table (business template).
TITLE_LINES = ["we will refer CR entries only", "summary of customers"]

#: Flat end-user sheet ("Conclusion"): one filter-friendly row per credit.
CONCLUSION_COLUMNS = ["Customer Name", "Date", "Time", "Document No.", "Mode", "Bank Name", "Amount"]

#: Payment-mode word patterns (whole-word, so 'PRODUCTS' never matches 'CTS').
_MODE_WORDS: list[tuple[str, str]] = [
    (r"\bRTGS\b", "RTGS"),
    (r"\bNEFT\b", "NEFT"),
    (r"\bIMPS\b", "IMPS"),
    (r"\bUPI\b", "UPI"),
    (r"\bCHEQUE\b", "Cheque"),
    (r"\bCHQ\b", "Cheque"),
    (r"\bCLG\b", "Cheque"),
    (r"\bCTS\b", "Cheque"),
    (r"\bCASH\b", "Cash"),
    (r"\bINB\b", "INB"),
]

#: First token of a slash-delimited IMPS/NEFT narration (MODE/ref/NAME/...).
_SLASH_MODES = {"IMPS", "NEFT", "RTGS", "UPI"}

#: Sheet names that ARE the bank label (upper-cased comparison).
_KNOWN_BANKS = {"SBI", "IDBI", "HDFC", "ICICI", "AXIS", "PNB", "BOB", "CANARA", "UNION", "KOTAK"}

#: Max unique narrations sent to the AI per request (keeps token cost bounded).
_AI_BATCH = 80

#: Company-suffix normalisation used for grouping keys (display keeps original).
_SUFFIX_MAP = [
    (r"\bPRIVATE\b", "PVT"),
    (r"\bLIMITED\b", "LTD"),
    (r"\bPVT\.\b", "PVT"),
    (r"\bLTD\.\b", "LTD"),
]


# ----- Small parsing helpers -------------------------------------------- #


def _norm(value: Any) -> str:
    """Normalize a header cell for tolerant matching (lowercase alnum only)."""
    return re.sub(r"[^a-z0-9]", "", str(value or "").lower())


def _num(value: Any) -> float | None:
    """Parse an Indian-formatted amount ('10,69,870.00', '168,740') to float."""
    if value is None:
        return None
    text = str(value).strip()
    if text in ("", "-", "--"):
        return None
    negative = text.startswith("(") and text.endswith(")")
    text = re.sub(r"[^0-9.]", "", text)
    if text in ("", "."):
        return None
    try:
        number = float(text)
    except ValueError:
        return None
    return -number if negative else number


def _fmt_date(value: Any) -> str:
    """Normalize a statement date to DD/MM/YYYY (best effort; raw on failure)."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    text = str(value).strip()
    if not text:
        return ""
    try:
        serial = float(text)
    except ValueError:
        serial = None
    if serial is not None and 20000 <= serial <= 80000:
        return (datetime(1899, 12, 30) + timedelta(days=int(serial))).strftime("%d/%m/%Y")
    text = text.split()[0]
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d-%b-%Y", "%d/%m/%y",
                "%d-%m-%y", "%d.%m.%Y", "%m/%d/%Y", "%m-%d-%Y"):
        try:
            return datetime.strptime(text, fmt).strftime("%d/%m/%Y")
        except ValueError:
            continue
    return text


def _date_sort_key(display: str) -> str:
    try:
        return datetime.strptime(display, "%d/%m/%Y").strftime("%Y%m%d")
    except ValueError:
        return ""


def _mode(description: Any) -> str:
    text = str(description or "").upper()
    if "BY BILL" in text or "LCBD" in text:
        return "Bill/LC"
    for pattern, label in _MODE_WORDS:
        if re.search(pattern, text):
            return label
    match = re.match(r"^([A-Z]{4})([RN])\d", text.strip())
    if match:
        return "RTGS" if match.group(2) == "R" else "NEFT"
    return ""


def _clean_name_tail(value: Any) -> str:
    text = re.sub(r"\s+", " ", str(value or "")).strip()
    return text.strip(" .-*/").strip()


def _letters_ratio(text: str) -> float:
    if not text:
        return 0.0
    return sum(c.isalpha() for c in text) / len(text)


def _looks_like_name(text: str) -> bool:
    cleaned = _clean_name_tail(text)
    return len(cleaned) >= 3 and _letters_ratio(cleaned.replace(" ", "")) >= 0.7


#: Narration segments that are remarks, never party names.
_JUNK_SEGMENTS = re.compile(
    r"^(BATCH?|BILL|BILLS|PAY|PYMT|PAYMENT|PAYMENTS|INV|INVOICE|ONLINE|OTHERS?|"
    r"TRANSFER|BANK\s*NOT?|IMPS|NEFT|RTGS|UPI)\b[-\s]*\S*$",
    re.IGNORECASE,
)


def _junk_segment(text: str) -> bool:
    return bool(_JUNK_SEGMENTS.match(_clean_name_tail(text)))


def _heuristic_name(description: Any, ref: Any = "") -> str:
    text = str(description or "").strip()
    reftext = re.sub(r"\s+", " ", str(ref or "")).strip()

    # 0) SBI ref column: 'TRANSFER FROM <acct> / <NAME>' or '<NAME> /'.
    if reftext:
        match = re.search(r"/\s*([A-Za-z][A-Za-z0-9 .&()\-]{2,})\s*$", reftext)
        if match and _looks_like_name(match.group(1)):
            return _clean_name_tail(match.group(1))
        match = re.search(r"TRANSFER FROM\s+\d+\s+([A-Za-z][A-Za-z .&()\-]{2,}?)\s*/",
                          reftext, re.IGNORECASE)
        if match and _looks_like_name(match.group(1)):
            return _clean_name_tail(match.group(1))

    if not text:
        return ""
    body = re.sub(r"^\s*BY TRANSFER[-\s]*", "", text, flags=re.IGNORECASE)

    # 1) '--NAME' suffix (SBI RTGS narration).
    match = re.search(r"--\s*([A-Za-z][A-Za-z0-9 .&()\-]{2,})\s*$", body)
    if match and _looks_like_name(match.group(1)):
        return _clean_name_tail(match.group(1))

    # 2) MODE/.../NAME/... (IMPS/UPI retail narration): first name-like field
    #    after the mode + reference ('IMPS/6144.../YASHIKA PR/...',
    #    'UPI/CR/6128.../SUDHAKAR/UTIB/...').
    if "/" in body:
        parts = [p.strip() for p in body.split("/")]
        if len(parts) >= 3 and parts[0].upper().split()[-1] in _SLASH_MODES:
            for part in parts[2:]:
                if _looks_like_name(part) and not _junk_segment(part):
                    return _clean_name_tail(part)

    # 3) '*'-delimited (SBI NEFT): the name is the segment AFTER the UTR —
    #    'NEFT*CBIN0282138*CBINH26138420737*NISHA KARKI*BATC--' -> 'NISHA KARKI'
    #    (the trailing segment is a remark like 'BATCH', never the name).
    if "*" in body:
        segments = [_clean_name_tail(s) for s in body.split("*")]
        last_ref = max((i for i, s in enumerate(segments)
                        if len(s) >= 10 and any(c.isdigit() for c in s)
                        and re.fullmatch(r"[A-Za-z0-9]+", s or " ")), default=0)
        candidates = [s for s in segments[last_ref + 1:]
                      if _looks_like_name(s) and not _junk_segment(s)]
        if not candidates:
            candidates = [s for s in segments[1:]
                          if _looks_like_name(s) and not _junk_segment(s)]
        if candidates:
            return candidates[0]

    # 4) 'NEFT-<ref>-NAME' (IDBI).
    match = re.match(r"^(?:NEFT|RTGS|IMPS|UPI)\s*-\s*[A-Za-z0-9]+\s*-\s*(.+)$",
                     body, re.IGNORECASE)
    if match and _looks_like_name(match.group(1)):
        return _clean_name_tail(match.group(1))

    # 5) Leading bank-reference token (has a digit, 10+ chars) then the name.
    match = re.match(r"^(?=[A-Za-z0-9]*\d)[A-Za-z0-9]{10,}\s+(.+)$", body)
    if match and _looks_like_name(match.group(1)):
        return _clean_name_tail(match.group(1))

    # 5b) Leading name then a long reference number ('INDIA GLYCOLS LTD 2601...').
    match = re.match(r"^([A-Za-z][A-Za-z .&()\-]{2,}?)\s+\d{6,}$", body)
    if match and _looks_like_name(match.group(1)):
        return _clean_name_tail(match.group(1))

    # 6) 'BULK POSTING-BY SALARY--' and similar.
    match = re.match(r"^BULK POSTING[-\s]*BY\s+(.+?)[-\s]*$", body, re.IGNORECASE)
    if match:
        return _clean_name_tail(match.group(1)).title()

    # 7) Cheque credits carry no payer name.
    if re.match(r"^CREDIT[-\s]*CHQ", body, re.IGNORECASE):
        return ""

    # 8) '-'-delimited fallback: trailing name-like segment.
    if "-" in body:
        tail = _clean_name_tail(body.split("-")[-1])
        if _looks_like_name(tail):
            return tail

    cleaned = _clean_name_tail(body)
    return cleaned if _looks_like_name(cleaned) else ""


def _utr_from_text(text: Any) -> str:
    for token in re.findall(r"[A-Z0-9]{10,}", str(text or "").upper()):
        if any(c.isdigit() for c in token) and any(c.isalpha() for c in token):
            return token
    tokens = re.findall(r"\d{10,}", str(text or ""))
    return tokens[0] if tokens else ""


def _doc_no(description: str, cheque_no: str, ref: str) -> str:
    desc = str(description or "")
    cheque = str(cheque_no or "").strip()
    if cheque and cheque not in ("0", "0.0", "nan"):
        return cheque

    match = re.search(r"UTR NO[:\s]*([A-Z0-9]{8,})", desc, re.IGNORECASE)
    if match:
        return match.group(1)
    match = re.search(r"\bCHQ\s*\.?\s*(\d{4,})", desc, re.IGNORECASE)
    if match:
        return match.group(1)
    match = re.search(r"(?:NEFT|RTGS|IMPS|UPI)[-/\s]*([A-Z0-9]{6,})", desc, re.IGNORECASE)
    if match and match.group(1).upper() not in ("TRANSFER",):
        return match.group(1)
    utr = _utr_from_text(desc)
    if utr:
        return utr

    reftext = str(ref or "")
    reftext = re.sub(r"TRANSFER\s+(FROM|TO)\s+\d+", " ", reftext, flags=re.IGNORECASE)
    match = re.search(r"([A-Z]{2,}[0-9][A-Z0-9]{4,})", reftext)
    if match:
        return match.group(1)
    token = _utr_from_text(reftext)
    return token


# ----- Table reading + bank/block detection ------------------------------ #


def _read_sheets(name: str, data: bytes) -> dict[str, Any]:
    import pandas as pd

    lower = name.lower()
    if lower.endswith(".csv"):
        import csv as _csv

        text = data.decode("utf-8-sig", errors="replace")
        rows = list(_csv.reader(io.StringIO(text)))
        width = max((len(r) for r in rows), default=0)
        rows = [r + [""] * (width - len(r)) for r in rows]
        return {"CSV": pd.DataFrame(rows, dtype=str).fillna("")}

    if lower.endswith(".xlsb"):
        engine = "pyxlsb"
    elif lower.endswith(".xls"):
        engine = "xlrd"
    else:
        engine = None
    frames = pd.read_excel(io.BytesIO(data), sheet_name=None, header=None,
                           dtype=str, engine=engine)
    return {str(k): v.fillna("") for k, v in frames.items()}


def _header_signature(cells: list[str]) -> str | None:
    cellset = {c for c in cells if c}
    if any(c.startswith("debit") for c in cellset) and any(c.startswith("credit") for c in cellset):
        return "SBI"
    if "crdr" in cellset and any(c.startswith("amount") for c in cellset):
        return "IDBI"
    return None


def _find_blocks(frame) -> list[tuple[int, str, dict[str, int]]]:
    blocks: list[tuple[int, str, dict[str, int]]] = []
    for i in range(len(frame)):
        cells = [_norm(x) for x in frame.iloc[i].tolist()]
        style = _header_signature(cells)
        if style is None:
            continue
        colmap: dict[str, int] = {}
        for pos, cell in enumerate(cells):
            if cell and cell not in colmap:
                colmap[cell] = pos
        blocks.append((i, style, colmap))
    return blocks


def _col(colmap: dict[str, int], *predicates) -> str | None:
    for predicate in predicates:
        for key in colmap:
            if (key == predicate) if isinstance(predicate, str) else predicate(key):
                return key
    return None


def _cell(row: list, colmap: dict[str, int], key: str | None) -> str:
    if key is None:
        return ""
    pos = colmap.get(key)
    if pos is None or pos >= len(row):
        return ""
    return str(row[pos]).strip()


def _parse_block(frame, start: int, end: int, style: str,
                 colmap: dict[str, int], bank: str) -> list[dict[str, Any]]:
    date_key = _col(colmap, "txndate", lambda k: "txndate" in k,
                    lambda k: k.startswith("date"), lambda k: "valuedate" in k)
    desc_key = _col(colmap, "description", lambda k: "descr" in k,
                    lambda k: "narrat" in k, lambda k: "particular" in k)
    if style == "IDBI":
        crdr_key = _col(colmap, "crdr")
        amount_key = _col(colmap, lambda k: k.startswith("amount"))
        cheque_key = _col(colmap, "chequeno", lambda k: "cheque" in k)
        ref_key = None
        balance_key = _col(colmap, lambda k: "balance" in k)
    else:  # SBI
        crdr_key = None
        amount_key = _col(colmap, "credit", lambda k: k.startswith("credit"))
        cheque_key = None
        ref_key = _col(colmap, lambda k: "refno" in k, lambda k: "cheque" in k)
        balance_key = _col(colmap, lambda k: "balance" in k)

    transactions: list[dict[str, Any]] = []
    for r in range(start, end):
        row = frame.iloc[r].tolist()
        if all(str(x).strip() == "" for x in row):
            continue
        if style == "IDBI":
            if _norm(_cell(row, colmap, crdr_key)) != "cr":
                continue
        amount = _num(_cell(row, colmap, amount_key))
        if amount is None or amount <= 0:
            continue
        raw_date = _cell(row, colmap, date_key)
        date_display = _fmt_date(raw_date)
        if not date_display:
            continue
        description = _cell(row, colmap, desc_key)
        cheque = _cell(row, colmap, cheque_key)
        ref = _cell(row, colmap, ref_key)
        balance = _num(_cell(row, colmap, balance_key))
        transactions.append({
            "date": date_display,
            "raw_date": raw_date,
            "description": description,
            "ref": ref,
            "amount": round(amount, 2),
            "balance": balance,
            "doc_no": _doc_no(description, cheque, ref),
            "mode": _mode(description),
            "bank": bank,
            "name_raw": _heuristic_name(description, ref),
        })
    return transactions


def _canon_datetime(value: Any) -> str:
    """Canonical 'YYYY-MM-DD HH:MM:SS' for a raw statement date cell.

    The same transaction is stored as '25/05/2026 04:08:54' in one pasted
    block and '2026-05-25 04:08:54.000' (or an Excel serial) in another —
    normalising the full timestamp lets de-duplication see them as equal.
    """
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    text = str(value or "").strip()
    if not text:
        return ""
    try:
        serial = float(text)
    except ValueError:
        serial = None
    if serial is not None and 20000 <= serial <= 80000:
        moment = datetime(1899, 12, 30) + timedelta(days=serial)
        return moment.strftime("%Y-%m-%d %H:%M:%S")
    text = re.sub(r"\.\d+$", "", text)  # drop fractional seconds
    for fmt in ("%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M:%S", "%d-%m-%Y %H:%M:%S",
                "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d-%b-%Y"):
        try:
            return datetime.strptime(text, fmt).strftime("%Y-%m-%d %H:%M:%S")
        except ValueError:
            continue
    return text


def _txn_key(txn: dict[str, Any]) -> tuple:
    """Identity of a transaction for de-duplication.

    Same posting timestamp + narration + reference + amount = the same
    transaction, however many times it was downloaded and pasted. The running
    balance is deliberately NOT part of the identity: a retroactive posting
    shifts the running balance of every later row between two downloads of the
    same account (observed in production data), which would fake uniqueness.
    Genuine repeat payments differ in timestamp or reference (UTR / cheque /
    LC-BD numbers live inside the narration).
    """
    return (
        _canon_datetime(txn.get("raw_date")) or txn["date"],
        re.sub(r"\s+", " ", txn["description"]).strip().upper(),
        re.sub(r"\s+", " ", str(txn.get("ref", ""))).strip().upper(),
        txn["amount"],
    )


def _dedupe_blocks(blocks_txns: list[list[dict[str, Any]]]) -> tuple[list[dict[str, Any]], int]:
    """Merge per-block transaction lists, dropping duplicate downloads.

    A transaction with the same canonical timestamp, narration, reference,
    amount and running balance as one already seen is the same bank posting
    (statements are downloaded repeatedly and pasted below each other, and
    sometimes the same range is pasted twice under one header) — keep one.

    Returns ``(unique_transactions, duplicates_removed)``.
    """
    seen: set = set()
    unique: list[dict[str, Any]] = []
    total = 0
    for txns in blocks_txns:
        for txn in txns:
            total += 1
            key = _txn_key(txn)
            if key in seen:
                continue
            seen.add(key)
            unique.append(txn)
    return unique, total - len(unique)


def _bank_label(sheet_name: str, style: str, filename: str) -> str:
    upper = re.sub(r"[^A-Z]", "", str(sheet_name or "").upper())
    if upper in _KNOWN_BANKS:
        return upper
    for bank in _KNOWN_BANKS:
        if re.search(rf"\b{bank}\b", str(filename or "").upper()):
            return bank
    return style


# ----- Hybrid AI name cleanup ------------------------------------------- #

_AI_SYSTEM = (
    "You normalise the paying customer/company name from Indian bank statement "
    "transaction narrations. Return ONLY strict JSON. Never invent a name that is "
    "not present in the narration."
)


def _clean_names_ai(transactions: list[dict[str, Any]], ai_gateway) -> dict[str, str]:
    unique: dict[str, str] = {}
    for txn in transactions:
        unique.setdefault(txn["description"], txn["name_raw"])
    items = list(unique.items())

    mapping: dict[str, str] = {}
    for start in range(0, len(items), _AI_BATCH):
        batch = items[start : start + _AI_BATCH]
        payload = [
            {"i": i, "narration": desc, "guess": guess}
            for i, (desc, guess) in enumerate(batch)
        ]
        instruction = (
            "For each item, extract the clean paying customer/party name from the "
            "bank narration. Rules: proper Title Case; drop bank/branch codes, "
            "IFSC, UTR/reference numbers, NEFT/RTGS/IMPS/UPI words, 'BY TRANSFER', "
            "batch/lot numbers and trailing dashes. If two items clearly name the "
            "same company, return the SAME canonical name for both. If unclear, "
            "return your best name from the narration (you may use 'guess').\n"
            "Respond as JSON exactly: {\"results\":[{\"i\":0,\"name\":\"...\"}]}\n"
            f"Items: {json.dumps(payload, ensure_ascii=False)}"
        )
        response = ai_gateway.extract(
            system_prompt=_AI_SYSTEM,
            instruction=instruction,
            parts=[],
            json_mode=True,
        )
        data = json.loads(response.text)
        for entry in data.get("results", []):
            idx = entry.get("i")
            name = str(entry.get("name") or "").strip()
            if isinstance(idx, int) and 0 <= idx < len(batch) and name:
                mapping[batch[idx][0]] = name
    return mapping


# ----- Customer grouping (canonical keys + truncated-name merge) --------- #


def _canonical(name: str) -> str:
    text = str(name or "").upper()
    for pattern, repl in _SUFFIX_MAP:
        text = re.sub(pattern, repl, text)
    return re.sub(r"[^A-Z0-9]", "", text)


def _merge_truncated(names: list[str]) -> dict[str, str]:
    by_key: dict[str, str] = {}
    for name in names:
        key = _canonical(name)
        if key and (key not in by_key or len(name) > len(by_key[key])):
            by_key[key] = name

    keys = sorted(by_key, key=len, reverse=True)
    resolve: dict[str, str] = {}
    for key in keys:
        if len(key) < 8:
            continue
        longer = [k for k in keys if k != key and k.startswith(key)]
        if len(longer) == 1:
            resolve[key] = resolve.get(longer[0], longer[0])

    mapping: dict[str, str] = {}
    for name in names:
        key = _canonical(name)
        final_key = resolve.get(key, key)
        mapping[name] = by_key.get(final_key, name)
    return mapping


# ----- Summary assembly + Excel ----------------------------------------- #


def _build_rows(transactions: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], float, int]:
    display = _merge_truncated([
        txn.get("clean_name") or txn.get("name_raw") or "(Unknown)"
        for txn in transactions
    ])
    groups: dict[str, list[dict[str, Any]]] = {}
    for txn in transactions:
        name = display.get(txn.get("clean_name") or txn.get("name_raw") or "(Unknown)",
                           "(Unknown)")
        groups.setdefault(name, []).append(txn)

    rows: list[dict[str, Any]] = []
    grand_total = 0.0
    for name in sorted(groups, key=lambda s: s.lower()):
        subtotal = 0.0
        entries = sorted(groups[name],
                         key=lambda t: (_date_sort_key(t.get("date", "")), t.get("amount", 0)))
        for txn in entries:
            amount = float(txn.get("amount") or 0.0)
            subtotal += amount
            rows.append({
                "Date": txn.get("date", ""),
                "Customer Name": name,
                "Amount": round(amount, 2),
                "Document No.": txn.get("doc_no", ""),
                "Mode": txn.get("mode", ""),
                "Bank Name": txn.get("bank", ""),
            })
        rows.append({
            "Date": "", "Customer Name": f"{name} Total", "Amount": round(subtotal, 2),
            "Document No.": "", "Mode": "", "Bank Name": "",
        })
        grand_total += subtotal

    if groups:
        rows.append({
            "Date": "", "Customer Name": "GRAND TOTAL", "Amount": round(grand_total, 2),
            "Document No.": "", "Mode": "", "Bank Name": "",
        })
    return rows, round(grand_total, 2), len(groups)


def _txn_time(txn: dict[str, Any]) -> str:
    """HH:MM:SS posting time from the raw statement date ('' when date-only)."""
    canon = _canon_datetime(txn.get("raw_date"))
    if len(canon) >= 19:
        time_part = canon[11:19]
        return "" if time_part == "00:00:00" else time_part
    return ""


def _build_conclusion_rows(transactions: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Flat rows for the 'Conclusion' sheet — one credit per row, no subtotal
    rows, sorted by customer A-Z then date/time (filter- and pivot-friendly)."""
    display = _merge_truncated([
        txn.get("clean_name") or txn.get("name_raw") or "(Unknown)"
        for txn in transactions
    ])
    decorated = []
    for txn in transactions:
        name = display.get(txn.get("clean_name") or txn.get("name_raw") or "(Unknown)",
                           "(Unknown)")
        decorated.append((name.lower(), _date_sort_key(txn.get("date", "")),
                          _txn_time(txn), txn, name))
    decorated.sort(key=lambda entry: entry[:3])
    rows = []
    for _, _, time_part, txn, name in decorated:
        rows.append({
            "Customer Name": name,
            "Date": txn.get("date", ""),
            "Time": time_part,
            "Document No.": txn.get("doc_no", ""),
            "Mode": txn.get("mode", ""),
            "Bank Name": txn.get("bank", ""),
            "Amount": round(float(txn.get("amount") or 0.0), 2),
        })
    return rows


def _write_conclusion_sheet(workbook, rows: list[dict[str, Any]]) -> None:
    """Render the flat 'Conclusion' sheet (plain header row, auto-filter on)."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    sheet = workbook.create_sheet(title="Conclusion")
    for col, header in enumerate(CONCLUSION_COLUMNS, start=1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for r, row in enumerate(rows, start=2):
        for col, header in enumerate(CONCLUSION_COLUMNS, start=1):
            cell = sheet.cell(row=r, column=col, value=row.get(header))
            cell.border = border
            if header == "Amount":
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right")

    widths = {"Customer Name": 38, "Date": 12, "Time": 10, "Document No.": 24,
              "Mode": 10, "Bank Name": 12, "Amount": 16}
    for col, header in enumerate(CONCLUSION_COLUMNS, start=1):
        sheet.column_dimensions[get_column_letter(col)].width = widths.get(header, 14)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:G{max(len(rows) + 1, 2)}"


def _is_total_row(row: dict[str, Any]) -> bool:
    name = str(row.get("Customer Name", ""))
    return name == "GRAND TOTAL" or name.endswith(" Total")


def _write_sheet(workbook, title: str, rows: list[dict[str, Any]]) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, italic=True, color="7F7F00")
    subtitle_font = Font(bold=True, size=12)
    total_font = Font(bold=True)
    total_fill = PatternFill("solid", fgColor="F2F2F2")
    grand_fill = PatternFill("solid", fgColor="DDEBF7")
    thin = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ncols = len(SUMMARY_COLUMNS)

    sheet = workbook.create_sheet(title=title[:31])

    sheet.cell(row=1, column=1, value=TITLE_LINES[0]).font = title_font
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    sheet.cell(row=2, column=1, value=TITLE_LINES[1]).font = subtitle_font
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)

    header_row = 3
    for col, header in enumerate(SUMMARY_COLUMNS, start=1):
        cell = sheet.cell(row=header_row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for r, row in enumerate(rows, start=header_row + 1):
        is_total = _is_total_row(row)
        is_grand = str(row.get("Customer Name", "")) == "GRAND TOTAL"
        for col, header in enumerate(SUMMARY_COLUMNS, start=1):
            cell = sheet.cell(row=r, column=col, value=row.get(header))
            cell.border = border
            if header == "Amount" and isinstance(row.get(header), (int, float)):
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right")
            if is_total:
                cell.font = total_font
                cell.fill = grand_fill if is_grand else total_fill

    widths = {"Date": 12, "Customer Name": 38, "Amount": 16,
              "Document No.": 24, "Mode": 10, "Bank Name": 12}
    for col, header in enumerate(SUMMARY_COLUMNS, start=1):
        sheet.column_dimensions[get_column_letter(col)].width = widths.get(header, 14)
    sheet.freeze_panes = f"A{header_row + 1}"


def build_summary_excel(per_bank_rows: dict[str, list[dict[str, Any]]],
                        combined_rows: list[dict[str, Any]],
                        conclusion_rows: list[dict[str, Any]] | None = None) -> bytes:
    """Workbook: 'Conclusion' (flat, end-user) first, then one sheet per bank,
    then 'All Banks' when several banks are present."""
    from openpyxl import Workbook

    workbook = Workbook()
    workbook.remove(workbook.active)
    if conclusion_rows:
        _write_conclusion_sheet(workbook, conclusion_rows)
    for bank in sorted(per_bank_rows):
        _write_sheet(workbook, bank, per_bank_rows[bank])
    if len(per_bank_rows) > 1:
        _write_sheet(workbook, "All Banks", combined_rows)
    if not workbook.sheetnames:
        _write_sheet(workbook, "Customer Summary", [])

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


# ----- Entry point ------------------------------------------------------- #


def run(files: list[tuple[str, bytes]], ai_gateway=None) -> dict[str, Any]:
    warnings: list[str] = []
    bank_blocks: dict[str, list[list[dict[str, Any]]]] = {}

    for name, data in files:
        try:
            sheets = _read_sheets(name, data)
        except Exception as exc:  # noqa: BLE001
            logger.exception("Failed reading %s", name)
            warnings.append(f"{name}: could not read file ({exc}).")
            continue

        recognised_any = False
        for sheet_name, frame in sheets.items():
            blocks = _find_blocks(frame)
            if not blocks:
                continue
            recognised_any = True
            boundaries = [b[0] for b in blocks] + [len(frame)]
            for index, (header_row, style, colmap) in enumerate(blocks):
                bank = _bank_label(sheet_name, style, name)
                txns = _parse_block(frame, header_row + 1, boundaries[index + 1],
                                    style, colmap, bank)
                if txns:
                    bank_blocks.setdefault(bank, []).append(txns)
        if not recognised_any:
            warnings.append(
                f"{name}: unrecognised format — expected SBI (Debit/Credit columns) "
                "or IDBI (CR/DR + Amount columns) on at least one sheet. Skipped."
            )

    transactions: list[dict[str, Any]] = []
    duplicates_removed = 0
    for bank in sorted(bank_blocks):
        unique, removed = _dedupe_blocks(bank_blocks[bank])
        duplicates_removed += removed
        transactions.extend(unique)
    if duplicates_removed:
        warnings.append(
            f"Removed {duplicates_removed} duplicate entries caused by overlapping "
            "statement downloads pasted into the same workbook."
        )

    ai_used = False
    if transactions:
        mapping: dict[str, str] = {}
        has_capacity = bool(ai_gateway is not None and getattr(ai_gateway, "has_capacity", lambda: False)())
        if has_capacity:
            try:
                mapping = _clean_names_ai(transactions, ai_gateway)
                ai_used = bool(mapping)
            except Exception as exc:  # noqa: BLE001
                logger.warning("AI name cleanup failed: %s", exc)
                warnings.append(f"AI name cleanup unavailable — used rule-based names. ({exc})")
        for txn in transactions:
            txn["clean_name"] = mapping.get(txn["description"]) or txn["name_raw"] or "(Unknown)"

    per_bank_rows: dict[str, list[dict[str, Any]]] = {}
    totals: dict[str, float] = {}
    for bank in sorted({t["bank"] for t in transactions}):
        bank_txns = [t for t in transactions if t["bank"] == bank]
        rows_bank, total_bank, _ = _build_rows(bank_txns)
        per_bank_rows[bank] = rows_bank
        totals[bank] = total_bank

    combined_rows, grand_total, customers = _build_rows(transactions)
    conclusion_rows = _build_conclusion_rows(transactions)

    stats = {
        "files": len(files),
        "credit_entries": len(transactions),
        "duplicates_removed": duplicates_removed,
        "customers": customers,
        "total_amount": grand_total,
        "per_bank_total": totals,
        "banks": sorted(per_bank_rows),
        "ai_used": ai_used,
    }

    return {
        "columns": SUMMARY_COLUMNS,
        "rows": combined_rows,
        "excel_bytes": build_summary_excel(per_bank_rows, combined_rows, conclusion_rows),
        "warnings": warnings,
        "stats": stats,
    }
