"""Generic, spec-driven per-document Excel exporter.

Provides a clean two-area workbook (a key/value "Summary" sheet and an optional
"Line Items" sheet) for any processor that does not ship its own
``exporter.py``. It is driven entirely by the processor's
:class:`~processors.spec.ProcessorSpec`, so a new processor gets a presentable
Excel export with zero custom code. Excel is always built from the normalized
JSON, never directly from AI output (see CLAUDE.md).
"""

from __future__ import annotations

import io
import logging
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from processors.spec import ProcessorSpec, get_path

logger = logging.getLogger(__name__)

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_SECTION_FONT = Font(bold=True, color="1F4E78", size=12)
_LABEL_FONT = Font(bold=True)
_THIN = Side(style="thin", color="BBBBBB")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def build_excel_bytes(data: dict[str, Any], spec: ProcessorSpec) -> bytes:
    """Render a per-document workbook to bytes for the given processor spec.

    Args:
        data: Normalized document data.
        spec: The processor specification (sections + line items).

    Returns:
        The workbook serialized as XLSX bytes.
    """
    workbook = Workbook()
    summary = workbook.active
    summary.title = (spec.document_type or "Summary")[:31]
    _build_summary_sheet(summary, data, spec)

    if spec.line_items_path:
        items_sheet = workbook.create_sheet("Line Items")
        _build_line_items_sheet(items_sheet, data, spec)

    buffer = io.BytesIO()
    workbook.save(buffer)
    logger.info("Generic workbook built for %s.", spec.use_case_key)
    return buffer.getvalue()


def _build_summary_sheet(sheet: Worksheet, data: dict[str, Any], spec: ProcessorSpec) -> None:
    """Populate the key/value summary sheet from the spec's sections."""
    row = 1
    for section in spec.sections:
        cell = sheet.cell(row=row, column=1, value=section.title)
        cell.font = _SECTION_FONT
        row += 1
        for fspec in section.fields:
            label_cell = sheet.cell(row=row, column=1, value=fspec.label)
            label_cell.font = _LABEL_FONT
            label_cell.alignment = Alignment(vertical="top")
            value_cell = sheet.cell(row=row, column=2, value=_displayable(get_path(data, fspec.path)))
            value_cell.alignment = Alignment(vertical="top", wrap_text=True)
            row += 1
        row += 1

    sheet.column_dimensions["A"].width = 26
    sheet.column_dimensions["B"].width = 70


def _build_line_items_sheet(sheet: Worksheet, data: dict[str, Any], spec: ProcessorSpec) -> None:
    """Populate the line-item table from the spec's column definitions."""
    items = get_path(data, spec.line_items_path) or []
    columns = spec.line_item_columns

    if not columns and items:
        # Fall back to keys present in the first row.
        keys = list(items[0].keys()) if isinstance(items[0], dict) else []
        for col_index, key in enumerate(keys, start=1):
            _header_cell(sheet, col_index, key)
        for row_index, item in enumerate(items, start=2):
            for col_index, key in enumerate(keys, start=1):
                _data_cell(sheet, row_index, col_index, (item or {}).get(key))
    else:
        for col_index, column in enumerate(columns, start=1):
            _header_cell(sheet, col_index, column.label)
        for row_index, item in enumerate(items, start=2):
            item = item or {}
            for col_index, column in enumerate(columns, start=1):
                _data_cell(sheet, row_index, col_index, item.get(column.key))

    _autosize(sheet)
    sheet.freeze_panes = "A2"


def _header_cell(sheet: Worksheet, col_index: int, text: str) -> None:
    """Write a styled header cell."""
    cell = sheet.cell(row=1, column=col_index, value=text)
    cell.fill = _HEADER_FILL
    cell.font = _HEADER_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = _BORDER


def _data_cell(sheet: Worksheet, row: int, col: int, value: Any) -> None:
    """Write a bordered data cell."""
    cell = sheet.cell(row=row, column=col, value=_displayable(value))
    cell.border = _BORDER
    cell.alignment = Alignment(vertical="top", wrap_text=True)


def _displayable(value: Any) -> Any:
    """Coerce a value into something openpyxl can write cleanly."""
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _autosize(sheet: Worksheet, max_width: int = 50) -> None:
    """Approximate column auto-sizing based on content length."""
    widths: dict[int, int] = {}
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            widths[cell.column] = max(widths.get(cell.column, 10), len(str(cell.value)) + 2)
    for column, width in widths.items():
        sheet.column_dimensions[get_column_letter(column)].width = min(width, max_width)
